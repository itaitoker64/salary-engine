"""
lookups.py — loads and exposes the grade (דרגה) and seniority (ותק) tables.

The grade table maps kod_darga (from the raw data) -> base salary at vatek=0.
The seniority table maps vatek (years) -> multiplier for grade type מנהלי (type 1).
"""

import openpyxl
from pathlib import Path


def load_lookups(excel_path: str) -> dict:
    """
    Load grade and seniority lookup tables from the גולמי Excel file.

    Returns a dict with:
      - 'grade':  {kod_darga: base_salary}
      - 'vatek':  {vatek_years: multiplier}
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # --- Grade table (דרגה sheet) ---
    # Columns: [idx, grade_label, internal_code, base_salary, pensionable, ...]
    # We build label -> base_salary first, then map kod_darga -> base from the raw data.
    ws_darga = wb["דרגה"]
    label_to_base: dict[str, float] = {}
    for row in ws_darga.iter_rows(min_row=1, values_only=True):
        label, base = row[1], row[3]
        if label is not None and base is not None:
            label_to_base[str(label)] = float(base)

    # kod_darga (e.g. 202, 210, 212) encodes both the grade type and whether
    # it's the plain or '+' sub-grade. We derive this mapping from the raw data
    # by observing which grade_label each kod_darga uses, then verifying with
    # the confirmed formula: base(label) × vatek_mult × job_pct = שכר משולב.
    #
    # Confirmed mapping (all grade type מנהלי = droog 1):
    # Plain grades (even tens):  160→'13', 170→'14', 180→'15', 190→'16',
    #                            200→'18', 210→'19', 220→'20', 230→'21',
    #                            240→'22', 250→'23', 260→'24'
    # Sub-grades (even tens+2):  172→'14', 182→'15', 192→'16', 196→'17',
    #                            202→'18', 212→'19+', 222→'20', 232→'21',
    #                            242→'22', 252→'23'
    #
    # The rule verified in data: kod_darga % 10 == 2 and kod % 20 == 12 → use '+' label.
    # Simpler: build the map directly from all observed pairs in the data.

    ws_golmi = wb["גולמי"]
    kod_to_label: dict[int, str] = {}
    for row in ws_golmi.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        # Only use the base salary rows (component 10002) to build the mapping
        if row[8] == 10002 and row[5] is not None and row[6] is not None:
            kod = int(row[5])
            label = str(row[6])
            if kod not in kod_to_label:
                kod_to_label[kod] = label

    # Build final grade lookup: kod_darga -> base_salary
    #
    # Rule confirmed from data: even tens (170, 180, 190...) → plain label ('18');
    # odd tens+2 (172, 182, 192...) → '+' label ('18+').
    # The pattern: if kod_darga % 20 == 2 → use label+; otherwise → use plain label.
    grade_lookup: dict[int, float] = {}
    for kod, label in kod_to_label.items():
        use_plus = (kod % 10 != 0)  # e.g. 172, 182, 192, 196, 202... → '+'; 170, 180, 190... → plain
        lookup_label = (label + '+') if use_plus else label
        base = label_to_base.get(lookup_label) or label_to_base.get(label)
        if base is not None:
            grade_lookup[kod] = base

    # --- Seniority table (ותק sheet) ---
    # Row 1: headers ('vetek', None, 'מנהלי')
    # Row 2: grade type codes (None, None, 1)
    # Row 3+: (vatek_years, sequential_code, multiplier_for_type_1)
    ws_vatek = wb["ותק"]
    vatek_lookup: dict[float, float] = {}
    for row in ws_vatek.iter_rows(min_row=3, values_only=True):
        vatek, _, mult = row[0], row[1], row[2]
        if vatek is not None and mult is not None:
            vatek_lookup[float(vatek)] = float(mult)

    wb.close()

    return {
        "grade": grade_lookup,
        "vatek": vatek_lookup,
        "label_to_base": label_to_base,
    }


def get_grade_base(lookups: dict, kod_darga: int) -> float | None:
    """Return base salary (at vatek=0) for a given kod_darga."""
    return lookups["grade"].get(int(kod_darga))


def get_vatek_multiplier(lookups: dict, vatek: float) -> float | None:
    """Return seniority multiplier for a given vatek (years). Exact match only."""
    return lookups["vatek"].get(float(vatek))
