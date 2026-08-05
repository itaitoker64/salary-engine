"""
main.py — Salary Engine API v0.2 (self-contained, flat structure)
"""

import os, io, re, sys, time, json, tempfile
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qsl, quote, urlencode
from dataclasses import dataclass, field
from collections import defaultdict, Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
import pandas as pd
from tools import progim_ingest  # top-level so Vercel bundles tools/*.py
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse, Response
from pydantic import BaseModel, Field

MATCH_THRESHOLD = 1.0
# Percentage tosafot ride the same base, which carries a small (~₪15) rounding
# "phantom" (see the 3.6% / מנמ"ש investigations): the identical base nudge that
# keeps the 3.6% gap under ₪1 pushes the 8–10% ones over it, so every worker
# with a 3.6% gap also shows one on מנמ"ש/2011/2024. Judging a percent component
# by its IMPLIED BASE (slip ÷ rate) instead of absolute ₪ clears all of a
# worker's base-noise gaps at once, while a real rate/level error (implied base
# off by hundreds) still fails. Threshold sits well above the noise cluster
# (≤₪23) and well below the genuine population (>₪100).
PERCENT_BASE_TOL = 25.0

# Default seniority track (קוד דרוג): 1 = מינהלי.
DEFAULT_TRACK = 1

# Base-salary component codes seen in the raw payroll dumps.
#   10002 = שכר משולב   — the combined base (base × seniority-multiplier × job%)
#       1 = יסוד משולב  — base at seniority 0 (× job%)
#       2 = תוספת ותק   — the seniority increment, base × (multiplier - 1) × job%
# The first form appears in older מנהלי files; the 1+2 split is used by the
# מנהלת הגמלאות (Pension Authority) dumps. Both reconstruct to base × mult × job%.
CODE_COMBINED_BASE = 10002
CODE_YESOD = 1
CODE_VETEK_TOSEFET = 2


def load_lookups(json_path: str) -> dict:
    """Load the multi-track lookup tables from lookups.json (built from Progim).

    Returns:
      - label_to_base:  {grade_label -> base salary at seniority 0}
      - vetek_by_track: {track_code -> {seniority_years -> multiplier}}
      - track_max:      {track_code -> max seniority years (cap)}
      - tracks:         {track_code -> track name}
    """
    raw = json.loads(Path(json_path).read_text(encoding="utf-8"))
    label_to_base = {str(k): float(v) for k, v in raw["darga"].items()}
    vetek_by_track = {
        int(track): {float(y): float(m) for y, m in table.items()}
        for track, table in raw["vetek"].items()
    }
    track_max = {int(k): float(v) for k, v in raw.get("track_max", {}).items()}
    tracks = {int(k): str(v) for k, v in raw.get("tracks", {}).items()}
    return {
        "label_to_base": label_to_base,
        "vetek_by_track": vetek_by_track,
        "track_max": track_max,
        "tracks": tracks,
    }


def normalize_grade_label(darga_label):
    """Canonicalize a grade label to the '<number>+' form used by the tables.

    Some dumps write intermediate grades with the '+' as a *prefix* ('+17')
    rather than a suffix ('17+'); older exports also pad with spaces. Move a
    leading '+' to the end so both spellings resolve to the same base."""
    if darga_label is None:
        return None
    s = str(darga_label).strip()
    if s.startswith("+") and s[1:].strip():
        s = s[1:].strip() + "+"
    return s


def get_grade_base(lookups, darga_label):
    """Base salary (seniority 0) for a grade label, e.g. '18', '42+'.

    Some גולמי dumps write intermediate grades with a leading plus ('+19',
    an RTL rendering artifact) while the pay table stores '19+' — normalize
    the prefix form to the suffix form before looking up."""
    if darga_label is None:
        return None
    return lookups["label_to_base"].get(normalize_grade_label(darga_label))


def get_vatek_multiplier(lookups, vatek, track=DEFAULT_TRACK):
    """Seniority multiplier for a given track and seniority (years).

    The multiplier is capped at the track's maximum seniority (e.g. מינהלי
    caps at 37 yrs, מח"ר at 40); beyond that the cap value is reused. The
    payroll table is on a 0.25-year grid, but real seniority is rarely on the
    grid — so for off-grid values the multiplier is **linearly interpolated**
    between the two surrounding grid points, which is what the source engine
    does (validated to ~93% exact base match on real data vs ~78% for nearest).
    """
    track = int(track)
    table = lookups["vetek_by_track"].get(track) or lookups["vetek_by_track"].get(DEFAULT_TRACK)
    if not table:
        return None
    vatek = float(vatek)
    cap = lookups["track_max"].get(track)
    if cap is not None:
        vatek = min(vatek, cap)
    if vatek in table:
        return table[vatek]
    keys = sorted(table)
    lower = [k for k in keys if k < vatek]
    upper = [k for k in keys if k > vatek]
    if not lower:
        return table[keys[0]]
    if not upper:
        return table[keys[-1]]
    a, b = lower[-1], upper[0]
    return table[a] + (table[b] - table[a]) * (vatek - a) / (b - a)

@dataclass
class WorkerInput:
    worker_id: int
    ministry_code: int
    ministry_name: str
    droog: int
    job_pct: float
    pension_pct: float
    kod_darga: int
    darga_label: str
    vatek_mandatory: float
    vatek_regular: float
    vatek_msc: float
    vatek_calculated: float
    calc_month: int
    retro_month: int
    retro_count: int
    components: list = field(default_factory=list)

@dataclass
class ComponentResult:
    code: int
    name: str
    amount: float
    pensionable: bool
    calculated: bool
    expected: Optional[float] = None
    diff: Optional[float] = None

@dataclass
class SalaryResult:
    worker_id: int
    ministry_code: int
    ministry_name: str
    droog: int
    kod_darga: int
    darga_label: str
    vatek_calculated: float
    job_pct: float
    pension_pct: float
    components: list
    total: float
    expected_total: Optional[float] = None
    total_diff: Optional[float] = None
    total_match: Optional[bool] = None
    status: str = "invalid"          # valid | invalid | no_base | multi_period
    grade_base: Optional[float] = None
    vatek_multiplier: Optional[float] = None
    errors: list = field(default_factory=list)

# Pay-slip classification statuses (Hebrew labels live in the frontend).
STATUS_VALID = "valid"            # תלוש תקין — computed base matches the slip
STATUS_INVALID = "invalid"        # תלוש שגוי — base present but does not match
STATUS_NO_BASE = "no_base"        # ללא שכר בסיס פעיל — no active base (pensioner/inactive)
STATUS_MULTI = "multi_period"     # רטרו / רב-תקופתי — multiple base periods on one slip

BASE_CODES = (CODE_COMBINED_BASE, CODE_YESOD, CODE_VETEK_TOSEFET)

# ---------------------------------------------------------------------------
# Component rules (החוקה) — extracted from the Progim workbook's SACHAR block by
# tools/extract_rules.py into component_rules.json:
#   percent — expected = rate × Σ(slip amounts of the rule's base codes); the
#             base composition comes from the actual חוקה formulas, and `rates`
#             lists every official rate (historical phase-ins / track variants).
#   manual  — components whose Netunei Gimlai column H says 'ידני': the value is
#             manually reported (Progim reads it from column J), so the engine
#             accepts the slip amount as-is and never flags it.
#
# Trust is SELF-CALIBRATING per file: a rule's mismatches are flagged only when
# the rule demonstrably holds for the file's own population (≥97% of the
# workers carrying the code match, with ≥20 carriers). A rule that fails
# file-wide (an era/track variation the model doesn't capture) is suppressed —
# a yellow cell must always mean a real, explainable gap.
# ---------------------------------------------------------------------------
TRUST_MIN_MATCH = 0.97
TRUST_MIN_N = 20


# The Progim covers PENSIONABLE pay only. These components are outside its
# scope by design — one-off or non-pensionable payments — so their absence from
# the workbook is not a coverage gap and they must not appear on the fix-list.
# Supplied by the user (the workbook's author); the גולמי "ביט פנסיוני" column
# cannot be used instead — on the 0108 file it reads 'כן' for all 124,818 rows,
# including for every code below, so it carries no signal.
NON_PENSIONABLE = {
    1927,   # ד. פגיעה בעבודה
    1936,   # תאונת עבודה
    1934,   # השלמה לפגיעה בעבודה
    1901,   # תגמול מילואים
    1266,   # דמי הבראה
    1260,   # דמי הבראה
    903,    # הפרש ברוטו
    889,    # הפרש ברוטו
    4457,   # הפקעת שכר
    1088,   # ימי שביתה
    1711,   # ניכוי 6% א"ע — ניכוי, לא תשלום (סכומו שלילי בגולמי)
    4120,   # השלמת שכר — מוזן ידנית מקובץ משרד האוצר, ואינו פנסיוני
    1375,   # קצובת ביגוד — קצובה, אינה פנסיונית
    4133,   # תוספת יוקר — אינה פנסיונית ואינה משתתפת בחישוב
    # Added 4.8.2026 on the user's instruction, off the 12/2010 coverage list.
    # None of the seven appears anywhere in the workbook (SACHAR / tosafot /
    # Netunei Gimlai / sminimum were all scanned), and none sits in any rule's
    # codes, base_codes or deductions — so declaring them out of scope only
    # states that their absence is by design. Three extend families already
    # here: 1269 joins 1266/1260 (דמי הבראה) and 4264 joins 903/889 (הפרש ברוטו).
    4962,   # מענק חד-פעמי — תשלום חד-פעמי
    4122,   # 100% ש.חופש — שווי חופשה
    4264,   # הפרש ברוטו — כמו 903/889
    4443,   # נסיעות — קצובה
    4121,   # 100% ש.זכ.שב — שווי זכאות שבתון
    5271,   # ימי חג
    1269,   # דמי הבראה — כמו 1266/1260
    # Added 4.8.2026 on the user's instruction, off the 12/2008 coverage list.
    # Same two checks as the batch above: absent from the workbook, absent from
    # every rule. 4535 does not appear in any sample file — declared anyway, so
    # that if it ever shows up it is already out of scope. Note 4536
    # (ממוצעי שכר) is 4537's twin and was NOT named; it stays in the gap list.
    1047,   # ימי עבודה
    4114,   # השתכרות חיצונית
    4537,   # ממוצעי שכר
    4535,   # ממוצעי שכר — לא נצפה באף קובץ דוגמה
    4454,   # מילואים-מל"ח
    4134,   # ת.יוקר-הפסק — כמו 4133 (תוספת יוקר)
    # Added 4.8.2026 on the user's instruction, off the 12/2011 coverage list.
    # Same two checks again: absent from the workbook, absent from every rule.
    5281,   # מענק שנתי — מענק, כמו 4962
    1265,   # דמי הבראה — הווריאנט הרביעי, אחרי 1266/1260/1269
    5272,   # ימי מחלה — היעדרות, לא רכיב שכר
    # Added 4.8.2026 on the user's instruction, off the 12/2014 coverage list.
    # 4123/4978 are absent from the workbook entirely. 1731 IS listed in
    # `sminimum` (row 188) — but only in the minimum-wage table, which declares
    # חוק מינימום / השלמה לשכר מינימום, not pensionability and not a formula.
    # Same precedent as 1375, already on this list and also in that table.
    1731,   # חתימה/עדות - גט
    4123,   # 100% ש.מחלה — היעדרות
    4978,   # נכ.העד-חרום
    1228,   # שט.יומית א' — קצובה יומית
    1229,   # שט.יומית ב' — קצובה יומית
    # Added 5.8.2026, off the 12/2016-12/2018 coverage lists. Both absent from
    # the workbook and from every rule. 5273 completes the 5271/5272 family.
    5438,   # סטודנט לומד
    5273,   # ימי בחירה — כמו 5271 (ימי חג) ו-5272 (ימי מחלה)
    4436,   # ימי חופשה — כמו 4122 (100% ש.חופש)
    4437,   # ימי מחלה — תאום של 5272
    5374,   # תוספת מו"מ
    # 4140 — added 5.8.2026 on the user's explicit instruction, AGAINST the
    # workbook. `מאפייני רכיבי שכר` declares 4140 "משכורת בסיסית שעתיים" with
    # **משכורת קובעת = כן**, i.e. pensionable. Declaring it out of scope
    # contradicts the Progim and removes ~₪90,200 (42 rows across 13 files)
    # from the coverage gap and from any check. The user was shown that flag
    # and chose to proceed. See docs/PROGIM_FIXES.md §17.
    4140,   # משכ. בסיסית — מוצהר פנסיוני בחוברת; הוצא בהוראת המשתמש
    # Added 5.8.2026 on the user's instruction. Unlike 4140 these AGREE with the
    # workbook: `מאפייני רכיבי שכר` declares 4192, 1152 and 1153 with
    # **משכורת קובעת = לא**. 651 has no entry there at all. None has a tosafot
    # or SACHAR column, and none appears in component_rules.json.
    4192,   # ביטוח מקיף - ידני (בגולמי: "ימים סגורים") — משכורת קובעת: לא
    1152,   # ש.נ. 125% — מוצהר משכורת קובעת: לא
    1153,   # ש.נ. 150% — מוצהר משכורת קובעת: לא
    651,    # מקדמה 93-96
}


def progim_coverage(rules) -> tuple:
    """Split the code universe by how the Progim treats each pay code.

    Returns (computable, referenced_only, reported):
      computable      — the Progim produces the amount from a formula/table
                        (percent, shekel, max22, minimum, the base split, gmul).
      reported        — type 'reported'/'manual': the Progim does NOT compute
                        the amount, it takes it as entered from the מנהלת
                        הגמלאות file. Accepting the slip value is therefore
                        CORRECT behaviour, not a coverage gap — but the amount
                        must still feed the bases of the percent components
                        that reference it (check_worker_components does that by
                        summing base_codes off the slip).
      referenced_only — the workbook knows the code only as an INPUT to another
                        formula (base member, 4550 deduction, minimum counted)
                        and has no entry of its own.

    A slip code in none of the three is unknown to the workbook entirely. Only
    the last two classes are real gaps in the product; the reports separate
    them from `reported` so the fix-list stays honest.
    """
    computable = {CODE_YESOD, CODE_VETEK_TOSEFET, CODE_COMBINED_BASE}
    computable |= set(GMUL_A_CODES) | set(GMUL_B_CODES)
    reported, referenced = set(), set()
    for rule in (rules or {}).values():
        target = reported if rule.get("type") in ("reported", "manual") else computable
        for c in rule.get("codes", []):
            target.add(int(c))
        for key in ("base_codes", "deductions", "counted", "toggle_codes"):
            for c in rule.get(key) or []:
                referenced.add(int(c))
    reported -= computable
    return computable, referenced - computable - reported, reported


def _grade_split_rates(gs, darga_label, droog):
    """Rates admissible for a grade-tiered tosefet (מנמ"ש 2010, code 5216).

    The rate steps down above a track-specific grade: מנהלי pays the higher
    level-1 rate up to 17+ and the lower level-2 rate from 18; מח"ר (track 11)
    steps at 38+/39. Each level keeps BOTH its historical and current rate
    (10%/7% at level 1, 8%/5% at level 2) so past pay periods still validate.
    Returns None when the grade can't be read — the caller then falls back to
    the full rate set, so a missing label never manufactures a false gap."""
    m = re.match(r"\s*\+?(\d+)", str(darga_label or ""))
    if not m:
        return None
    grade = int(m.group(1))
    try:
        is_machar = int(droog) == gs.get("machar_track")
    except (TypeError, ValueError):
        is_machar = False
    threshold = gs["machar_from_grade"] if is_machar else gs["default_from_grade"]
    return gs["level2_rates"] if grade >= threshold else gs["level1_rates"]


def check_worker_components(components, job_pct, rules, ministry_code=0,
                           darga_label=None, droog=None, pure: bool = False) -> dict:
    """Check each rule-covered percentage component on one slip.

    components: iterable of (code, name, amount, pensionable) slip rows.
    Returns {primary_code: {slip, expected, diff, ok, name}} for every percent
    rule whose component appears on the slip and whose base is computable.
    """
    amounts = defaultdict(float)
    for code, _name, amount, _pens in components:
        if code is not None:
            amounts[int(code)] += (amount or 0.0)
    checks = {}
    jp = job_pct or 1.0
    for code, rule in rules.items():
        rtype = rule["type"]
        if rtype not in ("percent", "shekel", "max22"):
            continue
        slip = sum(amounts.get(c, 0.0) for c in rule["codes"])
        if abs(slip) < 0.01:
            continue
        if rtype == "percent":
            base = sum(amounts.get(c, 0.0) for c in rule["base_codes"])
            base += rule.get("base_const", 0.0) * jp
            if base <= 0:
                continue
            # Grade-tiered tosefet (מנמ"ש 2010): restrict to the grade's own
            # level rates so a slip paid at the wrong level's rate is caught,
            # instead of silently accepting any of the four rates.
            rates = rule["rates"]
            # Per-grade rate table (נתיב: פו"מ 4319, תוספת שירות 4427). The
            # Progim resolves these with VLOOKUP(דרגה, <sheet>!C6:E115, 2, 0)
            # instead of a flat rate in `tosafot` row 3, so the grade decides
            # the rate outright. The workbook fills the table for grades 17–21
            # only; outside that range its VLOOKUP yields 0, which would mark
            # every payment wrong — so an unlisted grade skips the check and is
            # reported as a coverage gap instead of flagging the worker.
            rbg = rule.get("rate_by_grade")
            if rbg is not None:
                r = rbg.get(normalize_grade_label(darga_label))
                if r is None:
                    continue
                rates = [r]
            gs = rule.get("grade_split")
            if gs:
                lvl = _grade_split_rates(gs, darga_label, droog)
                if lvl:
                    rates = lvl
            best = min(rates, key=lambda r: abs(base * r - slip))
            expected = round(base * best, 2)
        elif rtype == "max22":
            # 4550 (הסכם 2001 אישי), per the Progim '4550' sheet: the higher of
            # 22% × (שכר משולב + הסכם 99) minus the listed deductions, and the
            # ministry floor (714.7 default; per-ministry overrides) × job%.
            base = sum(amounts.get(c, 0.0) for c in rule["base_codes"])
            if base <= 0:
                continue
            ded = sum(amounts.get(c, 0.0) for c in rule["deductions"])
            floor = rule["floors"].get(str(ministry_code or 0),
                                       rule["floor_default"]) * jp
            expected = round(max(rule["pct"] * base - ded, floor), 2)
        else:
            # shekel: the component is one of a fixed set of flat amounts, scaled
            # by job%. Expected = the closest admissible amount × job% (e.g.
            # גמול מינהל 4983 ∈ {105, 210, 315}). Wrong amounts still fail all.
            best = min(rule["amounts"], key=lambda a: abs(a * jp - slip))
            expected = round(best * jp, 2)
        ok = abs(expected - slip) <= MATCH_THRESHOLD
        if not pure and rtype == "percent" and not ok and best > 0:
            # base-relative pass: same base nudge that clears the 3.6% gap.
            # (An add-on beyond the literal Progim — off in pure mode.)
            ok = abs(slip / best - base) <= PERCENT_BASE_TOL
        checks[int(code)] = {
            "slip": round(slip, 2), "expected": expected,
            "diff": round(expected - slip, 2),
            "ok": ok,
            "name": rule["name"],
        }
    return checks


# גמולי השתלמות — checked against the file's own population, per the חוקה gate:
#   גמול א' (647/667/4268): one flat national amount (328.76 currently) — the
#     file-wide modal full-time value.
#   גמול ב' (897/4269): graded by kod_darga — lower grades a flat amount, higher
#     grades 9% × the base (SACHAR: IF(C2<47→P, C2<53→Q, else 9%)). The dominant
#     variant of each kod_darga group in the file IS the standard for that grade;
#     a deviation (e.g. a doubled or off-grid amount) usually means הפרשי רטרו
#     folded into the component.
GMUL_A_CODES = (647, 667, 4268)
GMUL_B_CODES = (897, 4269)
GMUL_A_NAME = "גמול השתלמות א'"
GMUL_B_NAME = "גמול השתלמות ב'"
GMUL_NOTE = "חריגה מהערך התקני לקבוצת הדרגה — חשד להפרשי רטרו בתוך הרכיב"


def check_gmul_population(entries) -> dict:
    """Self-calibrated gmul checks. Returns {entry_index: {code: flag}}."""
    a_vals, b_groups = Counter(), defaultdict(Counter)
    per = {}
    for i, e in enumerate(entries):
        r = e["result"]
        if r.status not in (STATUS_VALID, STATUS_INVALID):
            continue
        amt = defaultdict(float)
        for c in r.components:
            amt[c.code] += (c.expected or 0.0)
        a = sum(amt[c] for c in GMUL_A_CODES)
        b = sum(amt[c] for c in GMUL_B_CODES)
        base = sum(amt[c] for c in BASE_CODES)
        job = r.job_pct or 1.0
        per[i] = (a, b, base, job, r.kod_darga)
        if a > 0.01:
            a_vals[round(a / job, 2)] += 1
        if b > 0.01:
            if base > 1 and abs(b - 0.09 * base) <= MATCH_THRESHOLD:
                b_groups[r.kod_darga]["9%"] += 1
            else:
                b_groups[r.kod_darga][round(b / job, 2)] += 1
    a_mode, ta = None, sum(a_vals.values())
    if ta >= TRUST_MIN_N:
        v, n = a_vals.most_common(1)[0]
        if n / ta >= TRUST_MIN_MATCH:
            a_mode = v
    b_dom = {}
    for kod, c in b_groups.items():
        t = sum(c.values())
        if t >= 5:
            v, n = c.most_common(1)[0]
            if n / t >= 0.90:
                b_dom[kod] = v
    out = {}
    for i, (a, b, base, job, kod) in per.items():
        flags = {}
        if a_mode is not None and a > 0.01:
            exp = round(a_mode * job, 2)
            if abs(a - exp) > MATCH_THRESHOLD:
                flags[667] = {"slip": round(a, 2), "expected": exp,
                              "diff": round(exp - a, 2), "ok": False,
                              "name": GMUL_A_NAME, "note": GMUL_NOTE}
        dom = b_dom.get(kod)
        if dom is not None and b > 0.01:
            exp = round(0.09 * base, 2) if dom == "9%" else round(dom * job, 2)
            if abs(b - exp) > MATCH_THRESHOLD:
                flags[897] = {"slip": round(b, 2), "expected": exp,
                              "diff": round(exp - b, 2), "ok": False,
                              "name": GMUL_B_NAME, "note": GMUL_NOTE}
        if flags:
            out[i] = flags
    return out


# השלמות מינימום (1699 / 5260) — MAX(0, יעד − Σ רכיבים נספרים), where the
# counted set comes from the Progim sminimum sheet, the base counts at
# seniority 0 (נטרול ותק), and 4544's participation is a per-worker toggle.
# The minimum target itself is period-dependent, so it is inferred per file:
# the modal implied target (completion + counted sum, normalized to full
# time) across the file's own carriers. Tolerance is ₪3 — the expected value
# is a residual of a ~₪6,000 target minus a sum of many 2-decimal-rounded
# components, so ±₪1 would flag pure rounding accumulation.
MIN_TOLERANCE = 8.0


def check_minimum_population(entries, rules) -> dict:
    """Self-calibrated minimum-completion checks. {entry_index: {code: flag}}."""
    out = {}
    for code in (1699, 5260):
        rule = rules.get(code)
        if not rule or rule.get("type") != "minimum":
            continue
        counted = set(rule["counted"])
        toggles = rule.get("toggle_codes", [])
        data, cand = {}, Counter()
        for i, e in enumerate(entries):
            r = e["result"]
            if r.status not in (STATUS_VALID, STATUS_INVALID):
                continue
            amt = defaultdict(float)
            for c in r.components:
                amt[c.code] += (c.expected or 0.0)
            v = amt.get(code, 0.0)
            if v <= 0.01:
                continue
            job = r.job_pct or 1.0
            # Paid base for the minimum-wage sum: the slip reports it either as
            # יסוד משולב (1) or as the combined שכר משולב (10002). Use whichever
            # the slip actually carries — falling straight through to the grade
            # table would substitute the TABLE base for the PAID one and inflate
            # the implied target wherever the two differ (e.g. קוד-דרגה groups
            # whose label collides with a much higher מינהלי grade).
            yesod = amt.get(CODE_YESOD, 0.0) or amt.get(CODE_COMBINED_BASE, 0.0)
            if yesod <= 0:
                if r.grade_base is None:
                    continue
                yesod = r.grade_base * job
            csum = yesod + sum(amt.get(c, 0.0) for c in counted)
            tog = sum(amt.get(c, 0.0) for c in toggles)
            data[i] = (v, csum, tog, job)
            cand[round((v + csum) / job, 1)] += 1
            if tog:
                cand[round((v + csum + tog) / job, 1)] += 1
        if len(data) < TRUST_MIN_N or not cand:
            continue
        target = cand.most_common(1)[0][0]
        evals, ok_valid, n_valid = {}, 0, 0
        for i, (v, csum, tog, job) in data.items():
            e1 = max(0.0, round(target * job - csum, 2))
            e2 = max(0.0, round(target * job - csum - tog, 2))
            exp = e1 if abs(e1 - v) <= abs(e2 - v) else e2
            good = abs(exp - v) <= MIN_TOLERANCE
            evals[i] = (good, v, exp)
            # Trust is judged on the HEALTHY population: slips whose base
            # already validated. Broken slips (retro etc.) legitimately fail
            # this rule too and must not veto it.
            if entries[i]["result"].status == STATUS_VALID:
                n_valid += 1
                ok_valid += good
        if n_valid < TRUST_MIN_N or ok_valid / n_valid < TRUST_MIN_MATCH:
            continue  # era/model gap the file doesn't support — suppress
        for i, (good, v, exp) in evals.items():
            if good:
                continue
            out.setdefault(i, {})[code] = {
                "slip": round(v, 2), "expected": exp,
                "diff": round(exp - v, 2), "ok": False, "name": rule["name"],
                "note": (f"השלמת מינימום: היעד בקובץ ≈ {target} למשרה מלאה — "
                         "הסכום בתלוש אינו משלים אליו")}
    return out


def trusted_rule_codes(all_checks, rules=None) -> set:
    """Self-calibration: which rule codes hold on this file's own population.

    Rules marked "stable" in the חוקה (fixed rate that never phased, trivial
    base — e.g. ענ"א 16.2%/8%) are trusted even below the per-file sample
    threshold: with only a handful of carriers there is no population signal,
    but the rule itself is era-proof."""
    per_code = defaultdict(lambda: [0, 0])  # code -> [n, ok]
    for checks in all_checks:
        for code, chk in checks.items():
            per_code[code][0] += 1
            per_code[code][1] += chk["ok"]
    stable = {code for code, r in (rules or {}).items() if r.get("stable")}
    return {code for code, (n, ok) in per_code.items()
            if (n >= TRUST_MIN_N and ok / n >= TRUST_MIN_MATCH)
            or (code in stable and (n < TRUST_MIN_N or ok / n >= TRUST_MIN_MATCH))}


# The גולמי "ותק לחישוב שכר" column is a *rounded* seniority (to the nearest
# quarter-year, the resolution of the pay table), while the payroll engine used
# the exact, unrounded seniority. So a slip's base can legitimately differ from a
# recomputation off the rounded value by a couple of shekels. We therefore accept
# a base as correct when it is consistent with *any* seniority inside the rounding
# window (±0.125 yr) — this clears the ±₪1–2 precision artifacts without masking
# the real gaps, whose implied seniority is half a year or more off the stated ותק.
SENIORITY_ROUND = 0.125


def base_within_tolerance(grade_base, vatek, track, job_pct, slip_base, lookups):
    """Is slip_base consistent with the grade/track/job for some seniority within
    the ±0.125-yr rounding window of the stated ותק (with ±MATCH_THRESHOLD slack)?

    Returns True/False, or None when the multiplier can't be resolved (caller then
    falls back to the exact point comparison).
    """
    if grade_base is None:
        return None
    cap = lookups["track_max"].get(int(track))
    # The caller passes the truncation-corrected vetek (+0.005 when the file
    # value was not a whole quarter); widen the lower edge by the same amount
    # so the correction can only refine the check, never newly penalize.
    vlo, vhi = vatek - SENIORITY_ROUND - 0.005, vatek + SENIORITY_ROUND
    if cap is not None:
        vlo, vhi = min(vlo, cap), min(vhi, cap)
    m_lo = get_vatek_multiplier(lookups, vlo, track)
    m_hi = get_vatek_multiplier(lookups, vhi, track)
    if m_lo is None or m_hi is None:
        return None
    lo = min(m_lo, m_hi) * grade_base * (job_pct or 1.0)
    hi = max(m_lo, m_hi) * grade_base * (job_pct or 1.0)
    return (lo - MATCH_THRESHOLD) <= slip_base <= (hi + MATCH_THRESHOLD)


def normalize_vatek(v: float) -> float:
    """Undo the גולמי file's 2-decimal truncation of the seniority column.

    The payroll engine works on an eighth-of-year grid (…, 14.125, 14.375, …)
    but the גולמי export truncates to 2 decimals (14.125 → 14.12). Any value
    that is not a whole quarter (.0/.25/.5/.75) is therefore a truncated
    eighth — restore it by adding 0.005 (14.12 → 14.125)."""
    v = float(v)
    if round(v * 4, 6) % 1 != 0:
        return round(v + 0.005, 3)
    return v


def calculate(worker: WorkerInput, lookups: dict, pure: bool = False) -> SalaryResult:
    # pure=True runs the Progim literally: raw ותק (no truncation restore) and an
    # exact base match (no seniority-rounding window) — the app's data-cleaning
    # add-ons are off so the run reflects the workbook as-is.
    errors = []
    component_results = []
    total = 0.0
    track = int(worker.droog or DEFAULT_TRACK)
    if not pure:
        worker.vatek_calculated = normalize_vatek(worker.vatek_calculated or 0)
    grade_base = get_grade_base(lookups, worker.darga_label)
    if grade_base is None:
        errors.append(f"Unknown grade label: {worker.darga_label!r} (kod_darga {worker.kod_darga})")
    vatek_mult = get_vatek_multiplier(lookups, worker.vatek_calculated, track)
    if vatek_mult is None:
        errors.append(f"Unknown vatek/track: {worker.vatek_calculated}/{track}")
    job_pct = worker.job_pct or 1.0

    # Pre-scan the base components to classify the slip before computing:
    #  - raw base sum 0  → pensioner / inactive (no active base)
    #  - a primary base code (1 or 10002) appearing >1× → multiple periods / retro,
    #    which the single-period model cannot reconstruct from this file.
    raw_base_sum = sum((a or 0.0) for c, _, a, _ in worker.components if c in BASE_CODES)
    primary_base_count = max(
        sum(1 for c, *_ in worker.components if c == CODE_YESOD),
        sum(1 for c, *_ in worker.components if c == CODE_COMBINED_BASE),
    )
    if raw_base_sum <= MATCH_THRESHOLD:
        status = STATUS_NO_BASE
    elif primary_base_count > 1:
        status = STATUS_MULTI
    else:
        status = None  # decided after computing (valid/invalid)
    recompute = status is None  # only recompute base for active single-period slips

    for comp_code, comp_name, raw_amount, pensionable in worker.components:
        amount = raw_amount or 0.0
        calculated = False
        expected = raw_amount
        diff = None
        computed = None
        if recompute and grade_base is not None and vatek_mult is not None:
            if comp_code == CODE_COMBINED_BASE:
                # שכר משולב — full combined base
                computed = round(grade_base * vatek_mult * job_pct, 2)
            elif comp_code == CODE_YESOD:
                # יסוד משולב — base at seniority 0
                computed = round(grade_base * job_pct, 2)
            elif comp_code == CODE_VETEK_TOSEFET:
                # תוספת ותק — the seniority increment on top of יסוד
                computed = round(grade_base * (vatek_mult - 1.0) * job_pct, 2)
        if computed is not None:
            diff = round(computed - (raw_amount or 0.0), 4)
            amount = computed
            calculated = True
        total += amount
        component_results.append(ComponentResult(
            code=int(comp_code) if comp_code is not None else 0, name=comp_name or "", amount=amount,
            pensionable=(pensionable == "כן"), calculated=calculated,
            expected=expected, diff=diff,
        ))
    expected_total = sum((c[2] or 0.0) for c in worker.components)
    total_diff = round(total - expected_total, 4)
    total_match = abs(total_diff) <= MATCH_THRESHOLD
    if status is None:
        # Active single-period slip: judge the base against the seniority-rounding
        # window rather than the single rounded ותק value, so the ±₪1–2 artifacts
        # caused by the rounded ותק column aren't flagged as real errors.
        if pure:
            # Literal Progim: base = grade × ותק-multiplier × job%, matched to the
            # agora (MATCH_THRESHOLD), with no seniority-rounding window.
            if grade_base is not None and vatek_mult is not None:
                exp_base = grade_base * vatek_mult * job_pct
                total_match = abs(raw_base_sum - exp_base) <= MATCH_THRESHOLD
        else:
            base_ok = base_within_tolerance(
                grade_base, worker.vatek_calculated, track, job_pct, raw_base_sum, lookups)
            if base_ok is not None:
                total_match = base_ok
        if grade_base is None or vatek_mult is None:
            # The base could not be verified at all (unknown grade/track) — an
            # unverifiable active slip is never תקין by default.
            total_match = False
        status = STATUS_VALID if total_match else STATUS_INVALID
    # total_match only carries meaning for active single-period slips.
    if status in (STATUS_NO_BASE, STATUS_MULTI):
        total_match = None
    return SalaryResult(
        worker_id=worker.worker_id, ministry_code=worker.ministry_code,
        ministry_name=worker.ministry_name, droog=worker.droog,
        kod_darga=worker.kod_darga, darga_label=worker.darga_label,
        vatek_calculated=worker.vatek_calculated, job_pct=worker.job_pct,
        pension_pct=worker.pension_pct, components=component_results,
        total=round(total, 2), expected_total=round(expected_total, 2),
        total_diff=total_diff, total_match=total_match, status=status,
        grade_base=grade_base, vatek_multiplier=vatek_mult, errors=errors,
    )

# Maps a גולמי column header to a canonical field name. Matching is by Hebrew
# keyword (substring), so it is robust to column reordering and to the layout
# differences between the older מנהלי dumps and the מנהלת הגמלאות dumps (which
# carry two blank header rows and a different column order). Order matters:
# more specific patterns are checked first (e.g. "קוד דרגה" before "דרגה").
def _classify_header(h: str) -> Optional[str]:
    h = str(h).strip()
    # worker id — dumps label it מסד / מסב / מזהה, and newer exports use the
    # ID number instead ('תעודת זהות', 'ת.ז.', 'ת"ז', 'מספר זהות').
    _hz = h.replace('"', "").replace("'", "").replace(".", "").replace(" ", "")
    if ("מסד" in h or "מסב" in h or "מזהה" in h
            or ("מספר" in h and "עובד" in h)
            or "תעודתזהות" in _hz or "מספרזהות" in _hz or _hz in ("תז", "תזהות")):
        return "worker_id"
    if "קוד משרד" in h or "קוד גוף" in h or ("קוד" in h and "משרד/גוף" in h):
        return "ministry_code"
    if "שם משרד" in h or "שם גוף" in h or h in ("משרד/גוף", "משרד", "גוף"):
        return "ministry_name"
    if "דרוג" in h:
        return "droog"
    if "חלקיות" in h:
        return "job_pct"
    if "קוד דרגה" in h:
        return "kod_darga"
    if "קוד רכיב" in h:
        return "comp_code"
    if "רכיב" in h:
        return "comp_name"
    if "ותק" in h:
        return "vatek"
    if h == "דרגה" or ("דרגה" in h and "קוד" not in h):
        return "darga_label"
    if "פנסיו" in h:
        return "pensionable"
    if "סכום" in h or "סך" in h:
        return "amount"
    return None


def load_golmi(excel_path: str) -> dict:
    """Read a גולמי sheet, locate its header row, and group rows by worker.

    Columns are mapped by Hebrew header name rather than fixed position, so the
    same code reads both the old מנהלי layout (header on row 1) and the Pension
    Authority layout (two blank rows, header on row 3, different column order,
    base split into יסוד משולב + תוספת ותק).
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["גולמי"] if "גולמי" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Find the header row within the first few rows: the one that classifies the
    # most known fields (and at minimum has worker_id, a component and an amount).
    header_map, header_idx = None, None
    scan = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    best_score = 0
    for i, row in enumerate(scan):
        mapping = {}
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            field_name = _classify_header(cell)
            if field_name and field_name not in mapping:
                mapping[field_name] = ci
        score = len(mapping)
        if score > best_score and {"worker_id", "comp_code", "amount"} <= set(mapping):
            best_score, header_map, header_idx = score, mapping, i
    if header_map is None:
        wb.close()
        raise HTTPException(status_code=400,
                            detail="Could not find a recognizable גולמי header row in the file.")

    def cell(row, key):
        idx = header_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    workers = defaultdict(list)
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        wid = cell(row, "worker_id")
        if wid is None:
            continue
        workers[wid].append((
            cell(row, "ministry_code"), cell(row, "ministry_name"),
            cell(row, "droog"), cell(row, "job_pct"),
            cell(row, "kod_darga"), cell(row, "darga_label"), cell(row, "vatek"),
            cell(row, "comp_code"), cell(row, "comp_name"),
            cell(row, "pensionable"), cell(row, "amount") or 0.0,
        ))
    wb.close()
    return dict(workers)

# ---------------------------------------------------------------------------
# Plus-grade resolution — some גולמי dumps (the older מנהלי layout) drop the '+'
# from the דרגה label: a worker paid at grade '18+' is listed with דרגה '18',
# while קוד דרגה still distinguishes the plus grade (e.g. 200 = '18', 202 =
# '18+'). Left unresolved, every such worker looks ~2–7% off (the plus grades
# are exactly base × the grade's plus factor) and is falsely flagged invalid.
#
# The resolution is SELF-CALIBRATING per file, mirroring the rule-trust design:
# for each (קוד דרגה, label) population we reconstruct every active slip's base
# from the plain label and from its '+' variant, let the slips vote, and remap
# the code to the plus grade only when the vote is decisive (≥90% of deciding
# slips, ≥5 of them). Files whose labels already carry the '+' (the Pension
# Authority dumps) produce no votes for change and pass through untouched.
# ---------------------------------------------------------------------------
PLUS_MIN_VOTES = 5
PLUS_MIN_SHARE = 0.90


def resolve_plus_grades(workers_raw: dict, lookups: dict) -> dict:
    """Detect קוד-דרגה groups whose slips were actually paid at the '+' grade.

    Returns {(kod_darga, stated_label): resolved_label} for the groups where the
    file's own slips prove the label lost its '+'.
    """
    votes = defaultdict(lambda: [0, 0])  # (kod, label) -> [plain_fits, plus_fits]
    for _worker_id, rows in workers_raw.items():
        first = rows[0]
        _mc, _mn, droog, job_pct, kod_darga, darga_label, vatek = first[:7]
        label = str(darga_label or "").strip()
        if not label or label.endswith("+"):
            continue
        plus_base = get_grade_base(lookups, label + "+")
        plain_base = get_grade_base(lookups, label)
        if plus_base is None or plain_base is None:
            continue
        # Only active single-period slips can vote.
        slip_base = sum((r[10] or 0.0) for r in rows if r[7] in BASE_CODES)
        primary_count = max(sum(1 for r in rows if r[7] == CODE_YESOD),
                            sum(1 for r in rows if r[7] == CODE_COMBINED_BASE))
        if slip_base <= MATCH_THRESHOLD or primary_count > 1:
            continue
        track = int(droog or DEFAULT_TRACK)
        v = float(vatek or 0)
        pct = job_pct or 1.0
        plain_ok = base_within_tolerance(plain_base, v, track, pct, slip_base, lookups)
        plus_ok = base_within_tolerance(plus_base, v, track, pct, slip_base, lookups)
        if plain_ok and not plus_ok:
            votes[(kod_darga, label)][0] += 1
        elif plus_ok and not plain_ok:
            votes[(kod_darga, label)][1] += 1
    remap = {}
    for (kod, label), (plain_fits, plus_fits) in votes.items():
        decided = plain_fits + plus_fits
        if decided >= PLUS_MIN_VOTES and plus_fits / decided >= PLUS_MIN_SHARE:
            remap[(kod, label)] = label + "+"
    return remap


def run_engine_full(workers_raw: dict, lookups: dict, pure: bool = False) -> list:
    """Run the full engine over grouped גולמי rows: base validation per worker,
    then חוקה component checks with per-file self-calibration.

    Returns a list of dicts: {result, comp_checks, comp_flags} where comp_flags
    holds only the mismatches of rules that hold on this file's population —
    a flagged component is a real, explainable gap.

    pure=True runs the Progim workbook literally — no self-calibration trust
    gate, no base-relative tolerance, no plus-grade voting, no ותק restore, and
    no gmul/minimum population reconstruction. Every rule is applied as written
    and every mismatch is reported, so the run reflects the Progim as-is. The
    add-ons that would have cleared a gap are surfaced separately as
    recommendations (see build_progim_recommendations), not folded in silently.
    """
    rules = get_rules()
    entries = []
    # Pass A0 — resolve dropped-'+' grade labels from the file's own population.
    plus_remap = {} if pure else resolve_plus_grades(workers_raw, lookups)
    # Pass A — base validation + component checks per worker.
    for worker_id, rows in workers_raw.items():
        first = rows[0]
        ministry_code, ministry_name, droog, job_pct, kod_darga, darga_label, vatek = first[:7]
        darga_label = normalize_grade_label(darga_label) or darga_label
        darga_label = plus_remap.get((kod_darga, str(darga_label or "").strip()), darga_label)
        components = [(r[7], r[8], r[10], r[9]) for r in rows]
        worker = WorkerInput(
            worker_id=worker_id, ministry_code=ministry_code or 0,
            ministry_name=ministry_name or "", droog=droog or 1,
            job_pct=job_pct or 1.0, pension_pct=0.0,
            kod_darga=kod_darga or 0, darga_label=darga_label or "",
            vatek_mandatory=0.0, vatek_regular=float(vatek or 0),
            vatek_msc=0.0, vatek_calculated=float(vatek or 0),
            calc_month=0, retro_month=0, retro_count=0,
            components=components,
        )
        result = calculate(worker, lookups, pure=pure)
        # Component checks only make sense on active single-period slips.
        checks = (check_worker_components(components, worker.job_pct or 1.0, rules,
                                          worker.ministry_code,
                                          worker.darga_label, worker.droog, pure=pure)
                  if result.status in (STATUS_VALID, STATUS_INVALID) else {})
        # Base-only verdict, captured before pass B folds component flags in.
        entries.append({"result": result, "comp_checks": checks,
                        "base_ok": result.status == STATUS_VALID})
    # Pass B — attach flags. In pure mode every rule is trusted as written and
    # the gmul/minimum population reconstructions are skipped (they are add-ons,
    # not part of the literal Progim); otherwise self-calibrate rule trust.
    if pure:
        trusted = set(e_code for e in entries for e_code in e["comp_checks"])
        gmul_flags, min_flags = {}, {}
    else:
        trusted = trusted_rule_codes([e["comp_checks"] for e in entries], rules)
        gmul_flags = check_gmul_population(entries)
        min_flags = check_minimum_population(entries, rules)
    for i, e in enumerate(entries):
        flags = {code: chk for code, chk in e["comp_checks"].items()
                 if code in trusted and not chk["ok"]}
        flags.update(gmul_flags.get(i, {}))
        flags.update(min_flags.get(i, {}))
        e["comp_flags"] = flags
        result = e["result"]
        if flags:
            # A proven-wrong component makes the slip invalid even if its base
            # matches; record what is wrong and by how much.
            if result.status == STATUS_VALID:
                result.status = STATUS_INVALID
                result.total_match = False
            for code, chk in sorted(flags.items()):
                result.errors.append(
                    f"רכיב {code} ({chk['name']}): בתלוש {chk['slip']}, "
                    f"תקני {chk['expected']} (הפרש {chk['diff']})")
            # Fold the component corrections into the simulator total, so
            # "סכום מחושב" is the full corrected slip — base AND components —
            # and never equals the slip total when a component is wrong.
            comp_diff = round(sum(chk["diff"] for chk in flags.values()), 2)
            if comp_diff:
                result.total = round(result.total + comp_diff, 2)
                result.total_diff = round((result.total_diff or 0) + comp_diff, 4)
        e["findings"] = diagnose_entry(e, lookups)
    return entries


def build_progim_recommendations(pure_entries, smart_entries, rules) -> list:
    """Compare a pure-Progim run against the smart run and turn every gap the
    add-ons would have silently cleared into a Progim recommendation.

    We never patch the engine for these — the user decides what to add to the
    workbook. Returns a list of {category, code, name, count, sum, suggestion},
    most-frequent first, so each run says exactly what the Progim is missing.
    """
    smart_by_id = {e["result"].worker_id: e for e in smart_entries}
    recs = {}  # key -> {category, code, name, count, sum, suggestion}

    def bump(key, category, name, suggestion, amount=0.0, code=None):
        r = recs.setdefault(key, {"category": category, "code": code, "name": name,
                                  "count": 0, "sum": 0.0, "suggestion": suggestion})
        r["count"] += 1
        r["sum"] += abs(amount or 0.0)

    for pe in pure_entries:
        wid = pe["result"].worker_id
        se = smart_by_id.get(wid)
        if se is None:
            continue
        pr, sr = pe["result"], se["result"]
        # Base: rejected by the literal Progim but cleared by the smart layer's
        # ותק-rounding window / '+' resolution / truncation restore.
        if not pe.get("base_ok", True) and se.get("base_ok", False):
            bump("base_precision", "base_precision", "שכר בסיס",
                 "ה-Progim דורש ותק/תווית-דרגה מדויקים; הקובץ מעגל ותק ל-2 "
                 "ספרות ולעיתים משמיט '+' — לשקול לתקן במקור או להוסיף וריאנט",
                 round(pr.total_diff or 0.0, 2))
        # Components: any code flagged in pure but not in smart is a gap the
        # add-ons cleared — classify by why.
        for code, chk in pe["comp_flags"].items():
            if code in se["comp_flags"]:
                continue  # flagged in both — a genuine gap, not an add-on artifact
            name = chk.get("name", str(code))
            rtype = rules.get(code, {}).get("type")
            slip, exp = chk.get("slip", 0.0), chk.get("expected", 0.0)
            if rtype == "shekel":
                bump(("shekel", code), "shekel_mismatch", name,
                     f"הסכום השקלי של סמל {code} בחוקה אינו תואם את הקובץ — "
                     "לעדכן את טבלת הסכומים (דרגה×מסלול×פעימה) ב-Progim", chk.get("diff"), code)
            elif abs(round(exp - slip, 2)) <= 20.0:
                # small ₪ gap on a percent tosefet ⇒ the shared ~₪15 phantom base.
                bump(("base_noise", code), "base_noise", name,
                     "פער עיגול-בסיס קטן (‏~₪15 'בסיס-רפאים'); לשקול יישור הרכב "
                     "הבסיס בחוקה או קבלת אישור חשכ\"ל", chk.get("diff"), code)
            else:
                bump(("unstable_rule", code), "unstable_rule", name,
                     f"כלל {code} אינו מתאמת על מרבית נושאי הרכיב בקובץ — לבדוק "
                     "אחוז/בסיס בחוקה (ייתכן פעימה/סמל-בסיס חסר)", chk.get("diff"), code)

    out = sorted(recs.values(), key=lambda r: -r["count"])
    for r in out:
        r["sum"] = round(r["sum"], 2)
    return out


# ---------------------------------------------------------------------------
# דוח שגויים — itemized error diagnosis per non-valid slip.
#
# Every slip that is not תקין gets a precise, human-readable list of findings:
# what exactly is wrong, where, and by how much — e.g. תלוש ריק, בסיס כפול
# (2×, retro suspicion), פער ביסוד משולב, פער בתוספת ותק with the implied
# seniority the slip actually paid, or a gap in a specific חוקה component
# (גמול השתלמות, תוספת 3.6%...). The report ships as a "דוח שגויים" sheet in
# both exports.
# ---------------------------------------------------------------------------
BASE_NAMES = {CODE_YESOD: "יסוד משולב", CODE_COMBINED_BASE: "שכר משולב",
              CODE_VETEK_TOSEFET: "תוספת ותק"}
# Hebrew labels for the batch summary columns, in the DataFrame's own order —
# the Excel tabs must not show raw field names in a report that goes out.
# The English keys stay as-is: they are the CSV/API contract.
BATCH_HEADERS_HE = {
    "worker_id": "מסד עובד", "ministry_code": "קוד משרד",
    "ministry_name": "שם משרד", "droog": "דירוג", "kod_darga": "קוד דרגה",
    "darga_label": "דרגה", "vatek": "ותק", "job_pct": "חלקיות",
    "grade_base": "שכר יסוד (דרגה)", "vatek_mult": "מקדם ותק",
    "total_calculated": "סכום מחושב", "total_expected": "סכום בתלוש",
    "total_diff": "הפרש כולל", "gmul_diff": "הפרש גמולים",
    "total_match": "תואם", "status": "סטטוס",
    "flagged_components": "רכיבים חריגים", "n_components": "מס' רכיבים",
    "errors": "אבחון",
}

STATUS_HE = {STATUS_VALID: "תקין", STATUS_INVALID: "שגוי",
             STATUS_NO_BASE: "ללא שכר בסיס פעיל", STATUS_MULTI: "רטרו / רב-תקופתי"}


def implied_vatek(lookups, track, implied_mult):
    """The seniority (grid point) whose multiplier best matches implied_mult,
    or None when nothing on the track's grid comes close."""
    table = lookups["vetek_by_track"].get(int(track or DEFAULT_TRACK)) or {}
    if not table:
        return None
    best = min(table, key=lambda v: abs(table[v] - implied_mult))
    return best if abs(table[best] - implied_mult) <= 0.01 else None


def diagnose_entry(entry, lookups) -> list:
    """Itemized findings for one worker. Each finding:
    {category, code?, name?, slip?, expected?, diff?, note} — הפרש is always
    בתלוש minus תקני (positive = the slip pays more than the rulebook)."""
    result = entry["result"]
    findings = []
    if result.status == STATUS_NO_BASE:
        if abs(result.expected_total or 0) <= MATCH_THRESHOLD:
            findings.append({"category": "תלוש ריק",
                             "note": "אין סכומים בתלוש (סה\"כ ≈ 0)"})
        else:
            findings.append({"category": "ללא שכר בסיס פעיל",
                             "note": f"אין רכיב בסיס (יסוד/משולב) פעיל; "
                                     f"סה\"כ שאר הרכיבים {result.expected_total}"})
        return findings
    if result.status == STATUS_MULTI:
        findings.append({"category": "רטרו / רב-תקופתי",
                         "note": "רכיב בסיס ראשי מופיע יותר מפעם אחת — "
                                 "תלוש מרובה תקופות שכר"})
        return findings
    if result.status != STATUS_INVALID:
        return findings

    if result.grade_base is None:
        findings.append({"category": "דרגה לא מוכרת",
                         "note": f"דרגה '{result.darga_label}' לא נמצאה בטבלת השכר "
                                 f"(קוד דרגה {result.kod_darga})"})

    # Base-component gaps (יסוד משולב / תוספת ותק / שכר משולב).
    base_slip = base_exp = 0.0
    base_comps = {}
    for c in result.components:
        if c.calculated and c.code in BASE_NAMES:
            base_slip += (c.expected or 0.0)
            base_exp += c.amount
            base_comps[c.code] = c
    base_gap = round(base_slip - base_exp, 2)
    if abs(base_gap) > MATCH_THRESHOLD and base_exp > 0:
        ratio = base_slip / base_exp
        doubled = abs(ratio - 2.0) <= 0.01
        if doubled:
            findings.append({"category": "בסיס כפול (2×)",
                             "slip": round(base_slip, 2), "expected": round(base_exp, 2),
                             "diff": base_gap,
                             "note": "סכומי הבסיס בתלוש כפולים בדיוק מהתקן — "
                                     "חשד לתלוש רטרו / שתי תקופות מאוחדות"})
        for code in sorted(base_comps):
            c = base_comps[code]
            gap = round((c.expected or 0.0) - c.amount, 2)
            if abs(gap) > MATCH_THRESHOLD:
                findings.append({"category": f"פער ב{BASE_NAMES[code]}",
                                 "code": code, "name": BASE_NAMES[code],
                                 "slip": round(c.expected or 0.0, 2),
                                 "expected": round(c.amount, 2), "diff": gap})
        # What seniority does the slip's base actually correspond to?
        if result.grade_base and not doubled:
            implied = base_slip / (result.grade_base * (result.job_pct or 1.0))
            v = implied_vatek(lookups, result.droog, implied)
            if v is not None:
                years_gap = round(v - float(result.vatek_calculated or 0), 2)
                if abs(years_gap) >= 0.5:
                    findings.append({
                        "category": "ותק משתמע שונה מהרשום",
                        "note": f"הבסיס בתלוש תואם ותק של כ-{v} שנים "
                                f"(מקדם {round(implied, 4)}), לעומת "
                                f"{result.vatek_calculated} הרשום — פער "
                                f"{years_gap} שנים; ייתכן ותק-לתשלום שונה "
                                f"או שינוי דרגה במהלך החודש"})

    # חוקה component gaps (גמול השתלמות, תוספת 3.6%, הסכמי שכר...).
    for code, chk in sorted(entry.get("comp_flags", {}).items()):
        findings.append({"category": f"פער ב{chk['name']}",
                         "code": code, "name": chk["name"],
                         "slip": chk["slip"], "expected": chk["expected"],
                         "diff": round(chk["slip"] - chk["expected"], 2),
                         "note": chk.get("note",
                                         "לפי נוסחת החוקה (אחוז × סמלי הבסיס בתלוש)")})

    if not findings:
        findings.append({"category": "פער כולל",
                         "slip": result.expected_total, "expected": result.total,
                         "diff": round((result.expected_total or 0) - (result.total or 0), 2),
                         "note": "הסכום הכולל אינו תואם את החישוב"})
    return findings


REPORT_HEADERS = ["מספר עובד", "קוד משרד", "שם משרד", "דרגה", "ותק", "חלקיות",
                  "סטטוס", "קטגוריית שגיאה", "סמל", "שם רכיב",
                  "בתלוש", "תקני", "הפרש", "פירוט"]


def findings_report_rows(entries) -> list:
    """Flatten the per-worker findings into דוח-שגויים rows (non-valid only)."""
    rows = []
    for e in entries:
        r = e["result"]
        if r.status == STATUS_VALID:
            continue
        for f in e.get("findings", []):
            rows.append([
                r.worker_id, r.ministry_code, r.ministry_name, r.darga_label,
                r.vatek_calculated, r.job_pct, STATUS_HE.get(r.status, r.status),
                f.get("category"), f.get("code"), f.get("name"),
                f.get("slip"), f.get("expected"), f.get("diff"), f.get("note"),
            ])
    return rows


def append_report_sheet(wb, entries):
    """Add the דוח-שגויים sheet to a write-only workbook."""
    ws = wb.create_sheet("דוח שגויים")
    ws.sheet_view.rightToLeft = True
    header = []
    for h in REPORT_HEADERS:
        c = WriteOnlyCell(ws, value=h)
        c.font = Font(bold=True)
        header.append(c)
    ws.append(header)
    for row in findings_report_rows(entries):
        ws.append(row)
    return ws


def run_batch(excel_path: str, lookups: Optional[dict] = None) -> tuple:
    """Back-compat wrapper: (summary_df, detail_df) without the entries."""
    summary_df, detail_df, _entries = run_batch_entries(excel_path, lookups)
    return summary_df, detail_df


def run_batch_entries(excel_path: str, lookups: Optional[dict] = None) -> tuple:
    # Lookups come from the bundled engine data (lookups.json), not the uploaded
    # file — the Pension Authority dumps contain only raw rows, no lookup tables.
    if lookups is None:
        lookups = get_lookups()
    workers_raw = load_golmi(excel_path)
    summary_rows, detail_rows = [], []
    entries = run_engine_full(workers_raw, lookups)
    for entry in entries:
        result, flags = entry["result"], entry["comp_flags"]
        summary_rows.append({
            "worker_id": result.worker_id, "ministry_code": result.ministry_code,
            "ministry_name": result.ministry_name, "droog": result.droog,
            "kod_darga": result.kod_darga, "darga_label": result.darga_label,
            "vatek": result.vatek_calculated, "job_pct": result.job_pct,
            "grade_base": result.grade_base, "vatek_mult": result.vatek_multiplier,
            "total_calculated": result.total, "total_expected": result.expected_total,
            "total_diff": result.total_diff,
            "gmul_diff": (round(sum(chk["slip"] - chk["expected"]
                                    for code, chk in flags.items()
                                    if code in (667, 897)), 2) or None),
            "total_match": result.total_match,
            "status": result.status,
            "flagged_components": "; ".join(
                f"{code} ({chk['name']}): {chk['slip']} במקום {chk['expected']}"
                for code, chk in sorted(flags.items())),
            "n_components": len(result.components), "errors": "; ".join(result.errors),
        })
        for comp in result.components:
            detail_rows.append({
                "worker_id": result.worker_id, "ministry_code": result.ministry_code,
                "comp_code": comp.code, "comp_name": comp.name,
                "pensionable": comp.pensionable, "calculated": comp.calculated,
                "amount": comp.amount, "expected": comp.expected, "diff": comp.diff,
                "match": abs(comp.diff or 0) <= MATCH_THRESHOLD if comp.calculated else None,
            })
        for code, chk in sorted(flags.items()):
            detail_rows.append({
                "worker_id": result.worker_id, "ministry_code": result.ministry_code,
                "comp_code": code, "comp_name": chk["name"],
                "pensionable": None, "calculated": True,
                "amount": chk["expected"], "expected": chk["slip"],
                "diff": chk["diff"], "match": False,
            })
    return pd.DataFrame(summary_rows), pd.DataFrame(detail_rows), entries


# ---------------------------------------------------------------------------
# Highlighted export — "גולמי מעודכן" pivot with the invalid cells in yellow.
#
# The output keeps the exact layout of the מנהלת הגמלאות "גולמי מעודכן" sheet:
#   row 1–2  blank
#   row 3    component names (row label header in col A: "סכום של סכום")
#   row 4    field labels (קוד משרד … ותק) + the pay-code numbers + "סכום כולל"
#   row 5+   one row per worker — a column per pay code, plus the total in the
#            last column ("סכום כולל", i.e. column CO in the reference file).
#
# Only what is *not* valid is marked: every pay-code cell whose amount the engine
# could prove wrong is filled yellow, and the total cell is filled yellow when
# the slip's total is consequently off. Each highlighted cell carries a note with
# the value the engine expected, so the reviewer sees not only *that* it is wrong
# but *what it should be*. Valid slips, pensioners (no active base) and retro /
# multi-period slips are left untouched — the engine does not flag what it cannot
# prove, so a yellow cell always means a real, explainable gap.
# ---------------------------------------------------------------------------
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
META_LABELS = ["תוויות שורה", "קוד משרד", "שם משרד", "חלקיות משרה",
               "קוד דרגה", "דרגה", "ותק"]


def _num(v):
    """Coerce a raw cell to float, or None for blanks/non-numerics."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_highlighted_export(excel_path: str, lookups: Optional[dict] = None) -> io.BytesIO:
    """Rebuild the גולמי-מעודכן pivot and highlight only the invalid cells.

    Returns an in-memory .xlsx (BytesIO) ready to stream to the client.
    """
    if lookups is None:
        lookups = get_lookups()
    workers_raw = load_golmi(excel_path)

    per_worker = []        # rendered rows, in input order
    code_names = {}        # pay code -> display name (first one seen)
    all_codes = set()      # every pay code that appears anywhere in the file

    entries = run_engine_full(workers_raw, lookups)
    for (worker_id, rows), entry in zip(workers_raw.items(), entries):
        first = rows[0]
        ministry_code, ministry_name, droog, job_pct, kod_darga, darga_label, vatek = first[:7]
        components = [(r[7], r[8], r[10], r[9]) for r in rows]   # code, name, amount, pensionable
        result, comp_flags = entry["result"], entry["comp_flags"]

        # Pivot the slip: one summed amount per pay code (a code may repeat).
        slip_by_code = defaultdict(float)
        for code, name, amount, _pens in components:
            if code is None:
                continue
            code = int(code)
            slip_by_code[code] += (_num(amount) or 0.0)
            code_names.setdefault(code, name or str(code))
            all_codes.add(code)

        # Which pay codes are provably wrong, and what each one should have
        # been: the recomputed base when it mismatches, plus every חוקה
        # component whose (self-calibrated) rule fails on this slip.
        invalid_codes = {}
        if result.status == STATUS_INVALID:
            for comp in result.components:
                if comp.calculated and comp.diff is not None and abs(comp.diff) > MATCH_THRESHOLD:
                    invalid_codes[comp.code] = comp  # comp.amount = correct, comp.expected = slip
        # result.total already includes the חוקה component corrections
        # (folded in by run_engine_full) — do not add them again here.
        per_worker.append({
            "meta": [worker_id, ministry_code, ministry_name, job_pct,
                     kod_darga, darga_label, vatek],
            "slip_by_code": slip_by_code,
            "slip_total": result.expected_total,
            "corrected_total": round(result.total, 2),
            "invalid_codes": invalid_codes,
            "comp_flags": comp_flags,
            "total_invalid": result.status == STATUS_INVALID,
        })

    codes_sorted = sorted(all_codes)
    code_col = {code: len(META_LABELS) + i for i, code in enumerate(codes_sorted)}  # 0-based
    total_col = len(META_LABELS) + len(codes_sorted)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("גולמי מעודכן")
    ws.sheet_view.rightToLeft = True
    bold = Font(bold=True)

    # rows 1–2 blank (mirrors the reference file's two empty header rows)
    ws.append([])
    ws.append([])

    # row 3 — component names (col A carries the pivot's "סכום של סכום" label)
    row3 = ["סכום של סכום"] + [None] * (len(META_LABELS) - 1)
    row3 += [code_names.get(c, str(c)) for c in codes_sorted]
    row3 += [None]
    ws.append(row3)

    # row 4 — field labels + pay-code numbers + total label
    row4 = list(META_LABELS) + list(codes_sorted) + ["סכום כולל"]
    cells4 = []
    for v in row4:
        c = WriteOnlyCell(ws, value=v)
        c.font = bold
        cells4.append(c)
    ws.append(cells4)

    # rows 5+ — one per worker, invalid cells filled yellow with an expected-value note
    n_flagged = 0
    for w in per_worker:
        line = [None] * (total_col + 1)
        line[:len(META_LABELS)] = w["meta"]
        for code, amount in w["slip_by_code"].items():
            line[code_col[code]] = round(amount, 2)
        line[total_col] = round(w["slip_total"], 2) if w["slip_total"] is not None else None

        out = list(line)
        done_cols = set()
        for code, comp in w["invalid_codes"].items():
            ci = code_col[code]
            cell = WriteOnlyCell(ws, value=line[ci])
            cell.fill = YELLOW_FILL
            cell.comment = Comment(
                f"ערך תקני מחושב: {round(comp.amount, 2)}\n"
                f"בתלוש: {round(comp.expected or 0, 2)}\n"
                f"הפרש: {round((comp.expected or 0) - comp.amount, 2)}",
                "מנוע השכר")
            out[ci] = cell
            done_cols.add(ci)
            n_flagged += 1
        for code, chk in w["comp_flags"].items():
            ci = code_col.get(code)
            if ci is None or ci in done_cols:
                continue
            cell = WriteOnlyCell(ws, value=line[ci])
            cell.fill = YELLOW_FILL
            cell.comment = Comment(
                f"ערך תקני לפי החוקה: {chk['expected']}\n"
                f"בתלוש: {chk['slip']}\n"
                f"הפרש: {chk['diff']}",
                "מנוע השכר")
            out[ci] = cell
            done_cols.add(ci)
            n_flagged += 1
        if w["total_invalid"]:
            cell = WriteOnlyCell(ws, value=line[total_col])
            cell.fill = YELLOW_FILL
            cell.comment = Comment(
                f"סכום כולל מתוקן צפוי: {round(w['corrected_total'], 2)}\n"
                f"הפרש מהתלוש: {round(w['corrected_total'] - (w['slip_total'] or 0), 2)}",
                "מנוע השכר")
            out[total_col] = cell
            n_flagged += 1
        ws.append(out)

    append_report_sheet(wb, entries)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    n_invalid = sum(1 for w in per_worker if w["total_invalid"])
    return buf, len(per_worker), n_invalid, n_flagged


def content_disposition(hebrew_name: str, ascii_fallback: str) -> str:
    """Content-Disposition for a Hebrew download name. Browsers read the RFC-5987
    filename* (UTF-8); the plain filename= is an ASCII fallback for old clients."""
    return (f"attachment; filename=\"{ascii_fallback}\"; "
            f"filename*=UTF-8''{quote(hebrew_name)}")


app = FastAPI(title="Salary Engine API", version="0.2.0")


# --- the serverless rewrite ------------------------------------------------
# Vercel rewrites every path to this function, and hands it the rewrite's
# DESTINATION instead of the path the browser asked for. Every request arrived
# as the entrypoint path, no route matched, and the catch-all answered
# everything — `/api/lookups` included — with the frontend HTML: one identical
# 200 for every URL. That is what the page's "לא מחובר" badge really reports,
# since it fetches the pay tables and gets HTML it cannot parse.
#
# This has to live ON THE APP, not in api/index.py. A build wrapping the app
# there went live — proved by `/api/index` serving the page, which only the
# new main.py does — while not one response carried the wrapper's headers, so
# the runtime never calls the object that module exports.
#
# Nothing here assumes an undocumented header exists: each candidate is looked
# up by name, and `x-req-header-names` on the response reports what WAS sent,
# so the next step is a measurement rather than another guess. Names only —
# header values carry cookies and auth.
PATH_PARAM = "__path"
FORWARDED_PATH_HEADERS = (
    "x-vercel-original-path", "x-vercel-original-pathname",
    "x-original-uri", "x-forwarded-uri", "x-rewrite-url", "x-matched-path",
)


class _RestoreOriginalPath:
    """Put the browser's path back into the ASGI scope before routing."""

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope.get("type") != "http":
            await self.app(scope, receive, send)
            return

        received = scope.get("path", "")
        headers = {k.decode("latin-1").lower(): v.decode("latin-1")
                   for k, v in (scope.get("headers") or [])}
        pairs = parse_qsl((scope.get("query_string") or b"").decode("latin-1"),
                          keep_blank_values=True)

        original = next((v for k, v in pairs if k == PATH_PARAM), None)
        source = PATH_PARAM
        if not original:
            for name in FORWARDED_PATH_HEADERS:
                v = (headers.get(name) or "").split("?", 1)[0]
                if v.startswith("/") and v.strip("/") != received.strip("/"):
                    original, source = v, name
                    break

        if original:
            if not original.startswith("/"):
                original = "/" + original
            scope = dict(scope)
            scope["path"] = original
            scope["raw_path"] = original.encode("utf-8")
            scope["query_string"] = urlencode(
                [(k, v) for k, v in pairs if k != PATH_PARAM]).encode("latin-1")

        async def _send(message):
            if message.get("type") == "http.response.start":
                message = dict(message)
                message["headers"] = list(message.get("headers") or []) + [
                    (b"x-path-received", received.encode("latin-1", "replace")),
                    (b"x-path-restored",
                     (f"{original} via {source}" if original else "none")
                     .encode("latin-1", "replace")),
                    (b"x-req-header-names",
                     ",".join(sorted(headers)).encode("latin-1", "replace")[:900]),
                ]
            await send(message)

        await self.app(scope, receive, _send)


app.add_middleware(_RestoreOriginalPath)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BUNDLED_LOOKUPS = Path(__file__).parent / "lookups.json"
BUNDLED_RULES = Path(__file__).parent / "component_rules.json"
COMPONENTS_FILE = Path(__file__).parent / "components.json"
MINISTRIES_FILE = Path(__file__).parent / "ministries.json"
FRONTEND_FILE = Path(__file__).parent / "index.html"

# An uploaded Progim's extracted data lands here (writable at runtime). When a
# file exists it wins over the bundled copy, so an upload is live for this
# instance immediately; permanence across instances/deploys comes from the
# upload also being committed back to the repo (see /api/progim/upload).
PROGIM_DATA_DIR = Path(os.environ.get("PROGIM_DATA_DIR", "/tmp/progim_data"))

def _resolved(name: str, bundled: Path) -> Path:
    """The runtime-data copy of `name` if present, else the bundled file."""
    runtime = PROGIM_DATA_DIR / name
    return runtime if runtime.exists() else bundled

def LOOKUPS_FILE() -> Path:
    return _resolved("lookups.json", BUNDLED_LOOKUPS)

def RULES_FILE() -> Path:
    return _resolved("component_rules.json", BUNDLED_RULES)

_lookups: Optional[dict] = None
_rules: Optional[dict] = None
_components: Optional[dict] = None
_ministries: Optional[dict] = None

def _invalidate_data_caches():
    """Drop cached lookups/rules so the next access re-reads the (new) files."""
    global _lookups, _rules
    _lookups = None
    _rules = None

def get_rules() -> dict:
    """Component rules (החוקה) keyed by primary pay code (int)."""
    global _rules
    if _rules is None:
        rf = RULES_FILE()
        _rules = ({int(k): v for k, v in
                   json.loads(rf.read_text(encoding="utf-8")).items()}
                  if rf.exists() else {})
    return _rules

def get_components() -> dict:
    global _components
    if _components is None:
        _components = (json.loads(COMPONENTS_FILE.read_text(encoding="utf-8"))
                       if COMPONENTS_FILE.exists() else {"categories": {}, "components": []})
    return _components

def get_ministries() -> dict:
    global _ministries
    if _ministries is None:
        _ministries = (json.loads(MINISTRIES_FILE.read_text(encoding="utf-8"))
                       if MINISTRIES_FILE.exists() else {"ministries": []})
    return _ministries

def get_lookups() -> dict:
    global _lookups
    if _lookups is None:
        lf = LOOKUPS_FILE()
        if not lf.exists():
            raise RuntimeError(f"Lookup data file not found: {lf}")
        _lookups = load_lookups(str(lf))
    return _lookups

@app.on_event("startup")
async def startup():
    try:
        lk = get_lookups()
        print(f"Lookups loaded — grades: {len(lk['label_to_base'])}, "
              f"tracks: {len(lk['vetek_by_track'])}")
    except Exception as e:
        print(f"Failed to load lookups: {e}")

class ComponentInput(BaseModel):
    code: int
    name: str = ""
    amount: float = 0.0
    pensionable: str = "כן"

class CalculateRequest(BaseModel):
    worker_id: int = Field(..., example=11021106)
    ministry_code: int = Field(..., example=170)
    ministry_name: str = Field("", example="מכס ומע\"מ")
    droog: int = Field(1)
    job_pct: float = Field(1.0)
    pension_pct: float = Field(0.4)
    kod_darga: int = Field(..., example=202)
    darga_label: str = Field("", example="18")
    vatek_mandatory: float = Field(0.0)
    vatek_regular: float = Field(0.0)
    vatek_msc: float = Field(0.0)
    vatek_calculated: float = Field(..., example=33.75)
    calc_month: int = Field(228)
    retro_month: int = Field(0)
    retro_count: int = Field(1)
    components: list[ComponentInput] = Field(default_factory=list)

class ComponentOut(BaseModel):
    code: int; name: str; amount: float; pensionable: bool
    calculated: bool; expected: Optional[float]; diff: Optional[float]

class CalculateResponse(BaseModel):
    worker_id: int; ministry_code: int; ministry_name: str; droog: int
    kod_darga: int; darga_label: str; vatek_calculated: float
    job_pct: float; pension_pct: float; grade_base: Optional[float]
    vatek_multiplier: Optional[float]; components: list[ComponentOut]
    total: float; expected_total: Optional[float]
    total_diff: Optional[float]; total_match: Optional[bool]
    status: str; errors: list[str]

class AccuracyResponse(BaseModel):
    total_workers: int; matched: int; unmatched: int
    no_base: int = 0; multi_period: int = 0
    active_total: int = 0; active_accuracy_pct: float = 0.0
    accuracy_pct: float; match_threshold: float
    avg_diff: float; max_diff: float
    by_ministry: list[dict]; mismatches: list[dict] = []; elapsed_sec: float

def _frontend_response():
    """The single-page frontend. `no-cache` so a redeploy is picked up on the
    next load — the page carries an inline script whose BUILD must match the
    engine.js it asks for, and a stale HTML against a fresh engine.js is the
    one combination that silently mis-renders."""
    if FRONTEND_FILE.exists():
        return FileResponse(str(FRONTEND_FILE), media_type="text/html; charset=utf-8",
                            headers={"Cache-Control": "no-cache"})
    return JSONResponse({"status": "ok", "service": "salary-engine", "version": "0.2.0"})


@app.get("/", include_in_schema=False)
def root():
    return _frontend_response()

# Inline SVG favicon — navy rounded square with a white validation check, matching
# the frontend's <link rel="icon">. Served as a route so direct /favicon.ico hits
# (and the Vercel catch-all rewrite) don't 404.
FAVICON_SVG = (
    "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'>"
    "<rect width='32' height='32' rx='7' fill='#1E3A5F'/>"
    "<path d='M9 16.5l4.5 4.5L23 11' fill='none' stroke='#fff' stroke-width='3.2' "
    "stroke-linecap='round' stroke-linejoin='round'/></svg>"
)

@app.get("/favicon.ico", include_in_schema=False)
@app.get("/favicon.svg", include_in_schema=False)
def favicon():
    return Response(content=FAVICON_SVG, media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})

ENGINE_JS_FILE = Path(__file__).parent / "engine.js"

@app.get("/engine.js", include_in_schema=False)
def engine_js():
    """The validation engine, served to the browser so large גולמי files can be
    checked entirely client-side (no upload)."""
    return FileResponse(str(ENGINE_JS_FILE), media_type="application/javascript")

@app.get("/healthz")
def health():
    return {"status": "ok", "service": "salary-engine", "version": "0.2.0"}

@app.get("/api/info")
def info():
    # An unhandled raise here comes back as the platform's HTML error page,
    # which the frontend can only report as "לא מחובר" — the least useful
    # message possible. Answer with JSON either way, so the page can say what
    # actually broke.
    try:
        lk = get_lookups()
    except Exception as e:
        return JSONResponse(status_code=503, content={
            "status": "error", "error": f"{type(e).__name__}: {e}",
            "detail": "טבלאות השכר לא נטענו בשרת", "version": "0.3.0"})
    return {"status": "ok", "grades_loaded": len(lk["label_to_base"]),
            "tracks_loaded": len(lk["vetek_by_track"]),
            "track_caps": lk["track_max"],
            "match_threshold": MATCH_THRESHOLD, "version": "0.3.0"}


@app.get("/api/diag", include_in_schema=False)
def diag(request: Request):
    """What the function actually sees. This environment cannot reach the
    deployed site (network policy), so when production misbehaves this endpoint
    is the evidence: the path the platform handed us — a rewrite that passes its
    own destination instead of the request path is invisible from any other
    angle — and whether the data files survived the bundle. No secrets."""
    files = {}
    for p in (BUNDLED_LOOKUPS, BUNDLED_RULES, COMPONENTS_FILE, MINISTRIES_FILE,
              FRONTEND_FILE, ENGINE_JS_FILE):
        files[p.name] = p.stat().st_size if p.exists() else None
    try:
        lk = get_lookups()
        lookups = {"ok": True, "grades": len(lk["label_to_base"]),
                   "tracks": len(lk["vetek_by_track"])}
    except Exception as e:
        lookups = {"ok": False, "error": f"{type(e).__name__}: {e}"}
    try:
        rules_n = len(get_rules())
    except Exception as e:
        rules_n = f"{type(e).__name__}: {e}"
    return {
        "path_seen": request.url.path,
        "root_path": request.scope.get("root_path", ""),
        "python": sys.version.split()[0],
        "bundled_files": files,
        "runtime_data_dir": str(PROGIM_DATA_DIR),
        "runtime_data_present": PROGIM_DATA_DIR.exists(),
        "lookups": lookups,
        "rules_loaded": rules_n,
    }

@app.get("/api/lookups")
def api_lookups():
    """Full lookup tables (darga, vetek, track_max, tracks) as bundled in
    lookups.json. Small (~40 KB), so the browser can fetch them once and run the
    whole validation engine client-side — large גולמי files are then processed
    locally and never uploaded, sidestepping serverless request-body limits."""
    return json.loads(LOOKUPS_FILE().read_text(encoding="utf-8"))

@app.get("/api/rules")
def api_rules():
    """Component rules (החוקה) extracted from the Progim workbook — percentage
    bases/rates per pay code plus the manual (ידני) codes. Used by the browser
    engine for client-side component validation."""
    rf = RULES_FILE()
    return (json.loads(rf.read_text(encoding="utf-8")) if rf.exists() else {})


# ---------------------------------------------------------------------------
# עדכון ה-Progim מהאתר — admin uploads a new workbook; the engine data is
# regenerated (safe add-only merge) and persisted to PROGIM_DATA_DIR so every
# refresh, and every other computer hitting the same instance, uses the last
# uploaded file. Point PROGIM_DATA_DIR at a persistent disk (Render) for it to
# survive restarts and be shared across all clients. (An automated commit-back
# to the git repo — the only cross-instance option on Vercel's read-only FS — is
# deliberately NOT wired here: it would push to the protected branch at runtime,
# which must stay a reviewed action. See docs/PROGIM_UPDATE.md.)
# ---------------------------------------------------------------------------
def _persisted() -> bool:
    return (PROGIM_DATA_DIR / "component_rules.json").exists()


@app.get("/api/progim/status")
def progim_status():
    """Which Progim the engine is currently serving — the frontend shows the
    source and gates the upload UI on whether ADMIN_TOKEN is configured."""
    rf = RULES_FILE()
    n = len(json.loads(rf.read_text(encoding="utf-8"))) if rf.exists() else 0
    return {"source": "uploaded" if _persisted() else "bundled", "rules": n,
            "upload_enabled": bool(os.environ.get("ADMIN_TOKEN")),
            "data_dir": str(PROGIM_DATA_DIR)}


@app.post("/api/progim/upload")
async def progim_upload(file: UploadFile = File(...), token: str = Form("")):
    """Ingest an uploaded Progim (.xlsm/.xlsx): regenerate lookups + rules via a
    safe add-only merge and persist them to PROGIM_DATA_DIR, so subsequent loads
    (this instance) serve the uploaded file."""
    admin = os.environ.get("ADMIN_TOKEN")
    if not admin:
        raise HTTPException(503, "העלאת Progim מושבתת — יש להגדיר ADMIN_TOKEN בשרת")
    if token != admin:
        raise HTTPException(401, "סיסמת מנהל שגויה")
    raw = await file.read()
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(413, "הקובץ גדול מדי (מעל 25MB)")

    PROGIM_DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = PROGIM_DATA_DIR / "Progim_upload.xlsx"
    tmp.write_bytes(raw)

    # Ingest onto the CURRENT live rules (cumulative, add-only merge).
    cur = json.loads(RULES_FILE().read_text(encoding="utf-8"))
    try:
        lookups, rules, summary = progim_ingest.ingest(str(tmp), cur)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"קובץ ה-Progim לא נקרא/לא תקין: {str(e)[:200]}")
    if summary["rules"] < len(cur) or summary["grades"] < 50:
        raise HTTPException(400, "בדיקת שפיות נכשלה (ירידה במספר הכללים/דרגות) — "
                                 "כנראה קובץ שגוי; העדכון בוטל")

    (PROGIM_DATA_DIR / "lookups.json").write_text(
        json.dumps(lookups, ensure_ascii=False), encoding="utf-8")
    (PROGIM_DATA_DIR / "component_rules.json").write_text(
        json.dumps(rules, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    _invalidate_data_caches()

    persistent = str(PROGIM_DATA_DIR).startswith(("/var/", "/data", "/mnt"))
    note = ("נשמר בשרת ופעיל כעת. הקובץ ישמש כל ריענון/מחשב שמגיע לאותו מופע. "
            + ("השרת מוגדר עם דיסק קבוע — יישמר גם אחרי הפעלה מחדש."
               if persistent else
               "לשמירה קבועה בין הפעלות/מופעים — יש להגדיר דיסק קבוע (Render) "
               "או להטמיע את הקבצים ב-repo."))
    return {"ok": True, "filename": file.filename, "summary": summary, "note": note}

@app.get("/api/grades")
def list_grades():
    lk = get_lookups()
    return {"grades": [{"darga_label": k, "base_salary": v}
                       for k, v in sorted(lk["label_to_base"].items())]}

@app.get("/api/tracks")
def list_tracks():
    lk = get_lookups()
    return {"tracks": [{"code": k, "name": lk["tracks"].get(k, ""),
                        "max_vatek": lk["track_max"].get(k)}
                       for k in sorted(lk["vetek_by_track"])]}

@app.get("/api/ministries")
def list_ministries():
    """Ministries/units (code → name) that appear in the reference payroll data."""
    return get_ministries()

@app.get("/api/components")
def list_components():
    """Catalog of pay components: category, type, pensionable flag, and the
    typical contribution each makes (from the reference גולמי file)."""
    return get_components()

@app.get("/api/vatek/{years}")
def get_vatek(years: float, track: int = DEFAULT_TRACK):
    lk = get_lookups()
    mult = get_vatek_multiplier(lk, years, track)
    if mult is None:
        raise HTTPException(status_code=404, detail=f"No vatek entry for {years} years (track {track})")
    return {"vatek": years, "track": track, "multiplier": mult}

@app.post("/api/calculate", response_model=CalculateResponse)
def calculate_one(req: CalculateRequest):
    lk = get_lookups()
    worker = WorkerInput(
        worker_id=req.worker_id, ministry_code=req.ministry_code,
        ministry_name=req.ministry_name, droog=req.droog,
        job_pct=req.job_pct, pension_pct=req.pension_pct,
        kod_darga=req.kod_darga, darga_label=req.darga_label,
        vatek_mandatory=req.vatek_mandatory, vatek_regular=req.vatek_regular,
        vatek_msc=req.vatek_msc, vatek_calculated=req.vatek_calculated,
        calc_month=req.calc_month, retro_month=req.retro_month,
        retro_count=req.retro_count,
        components=[(c.code, c.name, c.amount, c.pensionable) for c in req.components],
    )
    result = calculate(worker, lk)
    return CalculateResponse(
        worker_id=result.worker_id, ministry_code=result.ministry_code,
        ministry_name=result.ministry_name, droog=result.droog,
        kod_darga=result.kod_darga, darga_label=result.darga_label,
        vatek_calculated=result.vatek_calculated, job_pct=result.job_pct,
        pension_pct=result.pension_pct, grade_base=result.grade_base,
        vatek_multiplier=result.vatek_multiplier,
        components=[ComponentOut(code=c.code, name=c.name, amount=c.amount,
            pensionable=c.pensionable, calculated=c.calculated,
            expected=c.expected, diff=c.diff) for c in result.components],
        total=result.total, expected_total=result.expected_total,
        total_diff=result.total_diff, total_match=result.total_match,
        status=result.status, errors=result.errors,
    )

@app.post("/api/accuracy", response_model=AccuracyResponse)
async def check_accuracy(file: UploadFile = File(...)):
    """Upload a גולמי Excel file. Returns accuracy % using ±1 ILS match threshold."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx")
    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content); tmp_path = tmp.name
    try:
        t0 = time.time()
        summary_df, detail_df = run_batch(tmp_path)
        elapsed = round(time.time() - t0, 1)
        total = len(summary_df)
        valid = int((summary_df["status"] == STATUS_VALID).sum())
        invalid = int((summary_df["status"] == STATUS_INVALID).sum())
        no_base = int((summary_df["status"] == STATUS_NO_BASE).sum())
        multi = int((summary_df["status"] == STATUS_MULTI).sum())
        # "Accuracy" is measured over the active single-period slips the model
        # actually validates (excluding pensioners and retro/multi-period rows).
        active = summary_df[summary_df["status"].isin([STATUS_VALID, STATUS_INVALID])]
        active_total = len(active)
        active_acc = round(valid / active_total * 100, 2) if active_total else 0.0
        overall_acc = round(valid / total * 100, 2) if total else 0.0
        diffs = active["total_diff"].abs()
        avg_diff = round(float(diffs.mean()), 4) if len(diffs) else 0.0
        max_diff = round(float(diffs.max()), 4) if len(diffs) else 0.0
        by_ministry = (
            active.groupby("ministry_name")
            .agg(workers=("worker_id", "count"),
                 matched=("status", lambda s: int((s == STATUS_VALID).sum())))
            .reset_index()
            .assign(accuracy_pct=lambda d: (d["matched"] / d["workers"] * 100).round(2))
            .sort_values("workers", ascending=False).head(20)
            .to_dict(orient="records")
        )
        # Per-worker gap detail for the invalid slips (largest gaps first), so the
        # UI can explain each mismatch: which base components differ and by how much.
        inv_df = (summary_df[summary_df["status"] == STATUS_INVALID]
                  .assign(absdiff=lambda d: d["total_diff"].abs())
                  .sort_values("absdiff", ascending=False).head(300))
        # Only the (≤300) invalid workers need per-component detail — grouping the
        # full 90k-row detail frame for every worker is needless time and memory.
        inv_ids = set(inv_df["worker_id"])
        calc_detail = detail_df[detail_df["calculated"] & detail_df["worker_id"].isin(inv_ids)]
        by_worker = {w: g for w, g in calc_detail.groupby("worker_id")}
        mismatches = []
        for _, r in inv_df.iterrows():
            comps = []
            for _, c in by_worker.get(r["worker_id"], pd.DataFrame()).iterrows():
                comps.append({
                    "code": int(c["comp_code"]), "name": c["comp_name"],
                    "slip": round(float(c["expected"] or 0), 2),
                    "computed": round(float(c["amount"] or 0), 2),
                    "diff": round(float(c["diff"] or 0), 2),
                })
            mismatches.append({
                "worker_id": int(r["worker_id"]), "ministry_name": r["ministry_name"],
                "darga_label": r["darga_label"], "vatek": r["vatek"], "job_pct": r["job_pct"],
                "grade_base": r["grade_base"], "vatek_multiplier": r["vatek_mult"],
                "total_calculated": r["total_calculated"], "total_expected": r["total_expected"],
                "total_diff": r["total_diff"], "components": comps,
            })
        return AccuracyResponse(
            total_workers=total, matched=valid, unmatched=invalid,
            no_base=no_base, multi_period=multi,
            active_total=active_total, active_accuracy_pct=active_acc,
            accuracy_pct=active_acc, match_threshold=MATCH_THRESHOLD,
            avg_diff=avg_diff, max_diff=max_diff,
            by_ministry=by_ministry, mismatches=mismatches, elapsed_sec=elapsed,
        )
    finally:
        os.unlink(tmp_path)

@app.post("/api/export-highlighted")
async def export_highlighted(file: UploadFile = File(...)):
    """Upload a גולמי Excel file → download the same גולמי-מעודכן pivot back,
    with every invalid pay-code amount and every wrong total marked in yellow."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx")
    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content); tmp_path = tmp.name
    try:
        t0 = time.time()
        buf, n_workers, n_invalid, n_flagged = build_highlighted_export(tmp_path)
        elapsed = round(time.time() - t0, 1)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": content_disposition(
                    "גולמי מסומן - שגיאות בצהוב.xlsx", "golmi_marked.xlsx"),
                "X-Workers": str(n_workers), "X-Invalid": str(n_invalid),
                "X-Cells-Flagged": str(n_flagged), "X-Elapsed-Sec": str(elapsed),
            },
        )
    finally:
        os.unlink(tmp_path)

@app.post("/api/batch")
async def batch_calculate(file: UploadFile = File(...)):
    """Upload a גולמי Excel file → download an .xlsx with three tabs: "תקין",
    "לבדיקה" (שגוי / ללא בסיס / רטרו), and "דוח שגויים" — an itemized error
    diagnosis per non-valid slip (what exactly is wrong, where, and by how much)."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx")
    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content); tmp_path = tmp.name
    try:
        t0 = time.time()
        summary_df, _, entries = run_batch_entries(tmp_path)
        elapsed = round(time.time() - t0, 1)
        valid_df = summary_df[summary_df["status"] == STATUS_VALID]
        review_df = summary_df[summary_df["status"] != STATUS_VALID]
        wb = Workbook(write_only=True)
        for sheet_name, df in (("תקין", valid_df), ("לבדיקה", review_df)):
            ws = wb.create_sheet(sheet_name)
            ws.sheet_view.rightToLeft = True
            cols = list(summary_df.columns)
            ws.append([BATCH_HEADERS_HE.get(c, c) for c in cols])
            i_match = cols.index("total_match") if "total_match" in cols else -1
            i_status = cols.index("status") if "status" in cols else -1
            clean = df.where(pd.notnull(df), None)
            for row in clean.itertuples(index=False):
                vals = [v.item() if hasattr(v, "item") else v for v in row]
                # Localize the two enum-ish cells too, so no English leaks out.
                if i_match >= 0 and isinstance(vals[i_match], bool):
                    vals[i_match] = "כן" if vals[i_match] else "לא"
                if i_status >= 0:
                    vals[i_status] = STATUS_HE.get(vals[i_status], vals[i_status])
                ws.append(vals)
        append_report_sheet(wb, entries)
        out = io.BytesIO(); wb.save(out); out.seek(0)
        total = len(summary_df); valid = len(valid_df)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": content_disposition(
                    "תוצאות בדיקת שכר - תקין ולבדיקה.xlsx", "salary_check_results.xlsx"),
                "X-Workers": str(total), "X-Valid": str(valid),
                "X-Review": str(total - valid), "X-Elapsed-Sec": str(elapsed),
            },
        )
    finally:
        os.unlink(tmp_path)


# ---------------------------------------------------------------------------
# Static files and the not-found fallback.
#
# On Vercel there is no static file server in front of us: vercel.json rewrites
# EVERY path to api/index.py, so FastAPI has to answer for the frontend's own
# files too. Anything without a route above returned Starlette's bare
# {"detail": "Not Found"} — which is what a user sees, and reads as "the site
# is down", not "wrong URL". Three real paths hit it:
#
#   /index.html, /salary_frontend.html  — a bookmark or a shared link
#   /api/index.py                       — a rewrite that hands the function its
#                                         destination path instead of the
#                                         original request path
#
# Registered last, so it only ever sees paths no route above claimed.
# ---------------------------------------------------------------------------
ROOT_DIR = Path(__file__).parent
# Both spellings: the rewrite destination may name the function's route
# ("api/index") or its file ("api/index.py"), and when the platform hands
# the function its destination instead of the requested path, THIS is the
# path that arrives. Recognising only one of them 404s the homepage.
VERCEL_ENTRY = {"api/index.py", "api/index"}
STATIC_SERVABLE = {
    "index.html": "text/html; charset=utf-8",
    "salary_frontend.html": "text/html; charset=utf-8",
    "engine.js": "application/javascript",
    "components.json": "application/json",
    "ministries.json": "application/json",
}


@app.get("/{full_path:path}", include_in_schema=False)
def static_or_frontend(full_path: str):
    path = full_path.strip("/")

    # A frontend file asked for by name — serve it from the repo root. Only
    # names on the list, and only from the root: no traversal, no data files.
    media = STATIC_SERVABLE.get(path)
    if media:
        f = ROOT_DIR / path
        if f.exists():
            headers = {"Cache-Control": "no-cache"} if path.endswith(".html") else {}
            return FileResponse(str(f), media_type=media, headers=headers)

    # The serverless entrypoint's own path — this is the app, so show the app.
    if path in VERCEL_ENTRY:
        return _frontend_response()

    # An unmatched /api/... path is a caller error, and an unmatched path that
    # looks like a file is a broken asset reference. Both deserve a real 404 —
    # answering them with HTML would turn a 404 into a confusing parse error.
    if path.startswith("api/") or "." in path.rsplit("/", 1)[-1]:
        raise HTTPException(status_code=404, detail="Not Found")

    # Anything else: a navigation. Show the app rather than a JSON error.
    return _frontend_response()
