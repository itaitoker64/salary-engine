# -*- coding: utf-8 -*-
"""
classify_hukka_amounts.py — split the חוקה amounts into fixed vs period-varying.

A component whose amount is a figure in the חוקה is one of two things, and the
distinction matters to whoever reads the report: either the workbook holds ONE
figure that has applied for the whole period, or it holds a **חודש פרישה** table
whose figure has actually changed over time. The second kind cannot be checked
against a single number, and a file from another month needs the figure of that
month — so the report must say which is which.

The decision is read out of the workbook, never assumed:

1. `tosafot` row 4 gives each component its column. Rows 7..234 of that column
   are the amount per חודש פרישה (row 7 = 1.2008), and row 3 is the current
   figure. Two or more distinct non-zero amounts down that column ⇒ **varies**.
2. When the column is empty the amount lives on a dedicated sheet, named by the
   formula in row 2/3. `SOURCE_GRIDS` maps each such sheet to the month × tier
   grid the formula points at. A tier column that holds one amount for all 228
   months is **fixed** — the spread is across tiers, not across time; two or
   more amounts down a single tier column is **varies**.
3. Anything that resolves to neither is left **unknown** and says so. Guessing
   "fixed" here would put an unsourced claim in the deliverable.

Usage:
    python3 tools/classify_hukka_amounts.py [Progim.xlsm] [component_rules.json]
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

TOSAFOT = "tosafot "
CODE_ROW = 4          # component code labels
CURRENT_ROW = 3       # the current figure (literal, or VLOOKUP into the block)
MONTH_ROW0, MONTH_ROW1 = 7, 234   # 1.2008 .. the last month the workbook carries

# Sheet -> the month x tier grid its tosafot formula points at.
#   cols/r0/r1 — the amount cells; rows are חודש פרישה.
#   tier        — what the COLUMNS are, read off the sheet's own header row.
#   tier_in_file— whether the גולמי identifies that group for a worker. This is
#                 a stated fact per sheet, not something inferred: getting it
#                 wrong either hides a checkable component or claims the file
#                 carries a field it does not.
#
# A component whose amount is chosen by a group the file does not carry cannot
# be validated at all — the workbook has the figure, we just cannot tell WHICH
# figure applies. Those get their own classification so they are not filed
# beside components a single number does check.
SOURCE_GRIDS = {
    "MANMASH 2022": {
        "cols": list("DEFG"), "r0": 5, "r1": 232,
        "tier": "דירוג (מינהלי / מח\"ר / טכנאים והנדסאים / מהנדסים)",
        # קוד דרוג is a column in the גולמי, so this one is resolvable in
        # principle — it is a coverage gap of ours, not an unanswerable one.
        "tier_in_file": True,
    },
    "5539 MEMUNE": {
        "cols": list("DEF"), "r0": 3, "r1": 230,
        "tier": "תעריף 1 / 2 / 3", "tier_in_file": False,
    },
    "5268 LIVUI ORHIM": {
        "cols": list("DEF"), "r0": 3, "r1": 230,
        "tier": "קבוצת ותק (עד 18 / 19–36 / 37+ חודשים)", "tier_in_file": False,
    },
    "tos reforma 4147": {
        "cols": [chr(c) for c in range(ord("H"), ord("S") + 1)], "r0": 7, "r1": 234,
        "tier": "תפקיד (ממונה / מנהל תחום / פקיד שומה / מפקח ...)",
        "tier_in_file": False,
    },
    # Pulse columns, not a worker group — the choice is by month alone.
    "heskem 2023": {"cols": list("GIKMO"), "r0": 4, "r1": 4, "tier": None,
                    "tier_in_file": True},
    "heskem 2016": {"cols": list("GIKMO"), "r0": 4, "r1": 4, "tier": None,
                    "tier_in_file": True},
}

SHEET_RE = re.compile(r"'([^']+)'!")


def _nums(values):
    return [round(float(x), 2) for x in values if isinstance(x, (int, float))]


def _column_of(tos_rows):
    """code -> 0-based column index, from the code labels in row 4."""
    out = {}
    for ci, label in enumerate(tos_rows[CODE_ROW - 1]):
        if label is None:
            continue
        for m in re.findall(r"\d{2,5}", str(label)):
            out.setdefault(int(m), ci)
    return out


def classify(wb_values, wb_formulas, code):
    """(period, evidence). period ∈ fixed | varies | group | unknown."""
    tos_v = list(wb_values[TOSAFOT].iter_rows(min_row=1, max_row=MONTH_ROW1,
                                              values_only=True))
    tos_f = wb_formulas[TOSAFOT]
    ci = _column_of(tos_v).get(code)
    if ci is None:
        return "unknown", "הסמל אינו מופיע בשורת הסמלים של tosafot"

    # Follow the formula to a source sheet FIRST. A component can carry a value
    # in its own tosafot column and still take its real amount from a group
    # grid — 5268 does, and deciding on the column alone filed it as a plain
    # fixed amount, hiding that the workbook cannot price it without a group
    # the file does not carry.
    from openpyxl.utils import get_column_letter
    col = get_column_letter(ci + 1)
    formula = " ".join(str(tos_f[f"{col}{r}"].value or "")
                       for r in (2, CURRENT_ROW))
    sheet = next((x for x in SHEET_RE.findall(formula) if x in SOURCE_GRIDS), None)

    months = _nums(tos_v[r][ci] for r in range(MONTH_ROW0 - 1, MONTH_ROW1))
    distinct = sorted(set(x for x in months if x))
    if sheet is None and len(distinct) >= 2:
        return "varies", (f"טבלת חודש פרישה ב-tosafot: {len(distinct)} סכומים "
                          f"שונים על פני {len(months)} חודשים "
                          f"({min(distinct):g}–{max(distinct):g})")
    if sheet is None and len(distinct) == 1:
        return "fixed", (f"טבלת חודש פרישה ב-tosafot: סכום אחד, {distinct[0]:g} ₪, "
                         f"לכל {len(months)} החודשים")

    if sheet is None:
        lit = tos_f[f"{col}{CURRENT_ROW}"].value
        if isinstance(lit, (int, float)):
            return "fixed", f"סכום קבוע בתא tosafot!{col}{CURRENT_ROW}: {float(lit):g} ₪"
        if isinstance(lit, str) and lit.startswith("=") and "$C$4" not in lit:
            # An IF/VLOOKUP on a worker attribute (דרגה, רמה, גיל) — the choice
            # is per worker, but the figures themselves do not move with time.
            return "fixed", ("סכום נבחר לפי מאפיין עובד (דרגה/רמה/גיל), "
                             "לא לפי חודש — קבוע לאורך התקופה")
        return "unknown", "עמודת החודשים ריקה ולא אותר גיליון מקור לסכום"

    grid = SOURCE_GRIDS[sheet]
    ws = wb_values[sheet]
    per_col = []
    for c in grid["cols"]:
        vals = _nums(ws[f"{c}{r}"].value for r in range(grid["r0"], grid["r1"] + 1))
        per_col.append(sorted(set(x for x in vals if x)))
    widest = max((len(d) for d in per_col), default=0)
    allv = sorted({x for d in per_col for x in d})

    # A group the file cannot identify makes the amount unresolvable no matter
    # how the months behave — that dominates, so it is decided first.
    if grid["tier"] and not grid["tier_in_file"] and sum(1 for d in per_col if d) > 1:
        moves = "ומשתנה גם לאורך התקופה" if widest >= 2 else "וקבוע לאורך התקופה"
        return "group", (f"גיליון '{sheet}': הסכום נבחר לפי {grid['tier']} {moves} "
                         f"({min(allv):g}–{max(allv):g} ₪). הקובץ הגולמי אינו נושא "
                         "את הקבוצה, ולכן לא ניתן לקבוע איזה סכום חל")
    if widest >= 2:
        return "varies", (f"גיליון '{sheet}' — טבלת חודש פרישה: עד {widest} סכומים "
                          f"שונים בתוך מדרג יחיד ({min(allv):g}–{max(allv):g})")
    if widest == 1:
        return "fixed", (f"גיליון '{sheet}': סכום אחד לכל חודש בכל מדרג "
                         f"({', '.join(f'{x:g}' for x in allv)} ₪ — ההבדל בין "
                         "המדרגים, לא בין החודשים)")
    return "unknown", f"גיליון '{sheet}' נמצא אך טבלת הסכומים ריקה"


def main():
    xlsm = Path(sys.argv[1] if len(sys.argv) > 1
                else "data/progim/Progim_31.07.2026.xlsm")
    rules_path = Path(sys.argv[2] if len(sys.argv) > 2 else "component_rules.json")
    wb_v = openpyxl.load_workbook(xlsm, data_only=True, keep_vba=True)
    wb_f = openpyxl.load_workbook(xlsm, data_only=False, keep_vba=True)
    rules = json.loads(rules_path.read_text(encoding="utf-8"))

    tally = {"fixed": 0, "varies": 0, "group": 0, "unknown": 0}
    for key, rule in rules.items():
        if rule.get("origin") != "hukka":
            rule.pop("amount_period", None)
            rule.pop("amount_period_note", None)
            continue
        code = int(rule["codes"][0])
        period, why = classify(wb_v, wb_f, code)
        rule["amount_period"] = period
        rule["amount_period_note"] = why
        tally[period] += 1
        print(f"{code:>6}  {period:<8} {why}")

    rules_path.write_text(json.dumps(rules, ensure_ascii=False, indent=2),
                          encoding="utf-8")
    print(f"\n{tally}  → {rules_path}")


if __name__ == "__main__":
    main()
