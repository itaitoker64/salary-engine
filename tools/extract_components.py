"""
extract_components.py — build components.json: a catalog of every pay component
(רכיב שכר) seen in a גולמי file, classified by what it contributes to the salary.

For each component code it records: name, how many workers carry it, its mean
amount, whether it is pensionable (משכורת קובעת), a category, a type
(percentage / fixed-shekel / base / deduction), and a short Hebrew explanation
of what the component gives. Output feeds the "how it works" tab.

Usage:
    python tools/extract_components.py path/to/golmi.xlsx components.json
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Category metadata: key -> (label, what it gives). Order = display priority.
CATEGORIES = {
    "base":      ("בסיס השכר", "ליבת השכר — נגזרת מהדרגה והוותק. כל שאר הרכיבים נבנים סביבה."),
    "minimum":   ("השלמה לשכר מינימום", "תוספת שמבטיחה שהשכר לא יורד מתחת לשכר המינימום החוקי."),
    "percent":   ("תוספות אחוזיות", "מחושבות כאחוז מהשכר המשולב — גדלות יחד עם הבסיס."),
    "shekel":    ("תוספות שקליות", "סכום קבוע בשקלים מכוח הסכמי שכר — אינו תלוי בגובה הבסיס."),
    "gmul":      ("גמולים והשתלמות", "תגמול על הכשרה/השתלמות או על תפקיד — לרוב אחוז מהבסיס."),
    "professional": ("תוספות מקצועיות", "תוספות לפי תפקיד/מקצוע (רפואה, מעבדה, כוננות) — מכוח הסכם ענפי."),
    "agreement": ("תוספות הסכם", "תוספות מכוח הסכמי שכר קיבוציים לפי שנה."),
    "deduction": ("ניכויים", "רכיבים שליליים — היעדרויות, ימי חופשה ללא תשלום וכד'."),
    "other":     ("רכיבים נוספים", "תוספות והפרשים נוספים."),
}


def classify(code: int, name: str, mean: float) -> tuple[str, str]:
    """Return (category_key, type_label) for a component."""
    n = name or ""
    if code in (1, 2, 10002):
        return "base", "בסיס"
    if mean < 0 or "העדרות" in n or "ניכוי" in n:
        return "deduction", "ניכוי"
    if "מינימום" in n:
        return "minimum", "השלמה"
    if "השתל" in n or "גמול" in n:
        return "gmul", "אחוז/סכום"
    if "אחוזית" in n or "%" in n:
        return "percent", "אחוז מהבסיס"
    if "שקלית" in n:
        return "shekel", "סכום קבוע"
    if any(k in n for k in ("בי\"ח", 'בי"ח', "קלינ", "פסיכ", "דריכות", "רפוא", "מעבד", "כונן")):
        return "professional", "תוספת מקצועית"
    if "הסכם" in n:
        return "agreement", "סכום/אחוז"
    if "מנמ" in n:  # מנמ"ש — administrative-professional agreement
        return "agreement", "סכום קבוע"
    if "השלמת שכר" in n:
        return "minimum", "השלמה"
    return "other", "תוספת"


def extract(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["גולמי"] if "גולמי" in wb.sheetnames else wb[wb.sheetnames[0]]
    # find header row (handles the 2 blank rows of the Pension Authority layout)
    rows = list(ws.iter_rows(values_only=True))
    header_i = next((i for i, r in enumerate(rows[:8])
                     if r and any(str(c).strip() == "קוד רכיב שכר" for c in r if c)), 0)
    header = [str(c).strip() if c is not None else "" for c in rows[header_i]]
    def col(name):
        return header.index(name) if name in header else None
    ci_code, ci_name = col("קוד רכיב שכר"), col("רכיב שכר")
    ci_pens, ci_amt = col("ביט פנסיוני"), col("סכום")

    agg = defaultdict(lambda: {"name": "", "n": 0, "sum": 0.0, "pens": 0})
    for r in rows[header_i + 1:]:
        if ci_code is None or ci_code >= len(r) or r[ci_code] is None:
            continue
        code = int(r[ci_code])
        a = agg[code]
        a["name"] = (r[ci_name] if ci_name is not None and r[ci_name] else a["name"]) or ""
        a["n"] += 1
        a["sum"] += float(r[ci_amt] or 0) if ci_amt is not None else 0.0
        if ci_pens is not None and r[ci_pens] == "כן":
            a["pens"] += 1
    wb.close()

    comps = []
    for code, a in agg.items():
        mean = round(a["sum"] / a["n"], 2) if a["n"] else 0.0
        cat, typ = classify(code, a["name"], mean)
        comps.append({
            "code": code, "name": a["name"], "n": a["n"], "mean": mean,
            "pensionable": a["pens"] > a["n"] / 2, "category": cat, "type": typ,
        })
    comps.sort(key=lambda c: -abs(c["mean"] * c["n"]))
    return {"categories": CATEGORIES, "components": comps}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    out = sys.argv[2] if len(sys.argv) > 2 else "components.json"
    data = extract(sys.argv[1])
    Path(out).write_text(json.dumps(data, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"Wrote {out}: {len(data['components'])} components in {len(data['categories'])} categories")
