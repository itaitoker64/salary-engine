# -*- coding: utf-8 -*-
"""
unified_report.py — run the engine on one or more monthly גולמי files and write
a polished, workable unified workbook:

  לוח בקרה     — KPI tiles + per-file summary with conditional formatting
  שגויים לבדיקה — the work queue: only invalid slips, sorted by |gap|, filterable
  פר עובד      — every employee from every file as a filterable Excel table
  פילוח משרדים — ministry × validity aggregated across all files

Usage:
    python tools/unified_report.py file1.xlsx [file2.xlsx ...] --out unified.xlsx
"""

import argparse
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import main as engine

# ---- palette (light surface; status colors paired with a text label, never alone)
NAVY = "FF1E3A5F"          # brand header
NAVY_TEXT = "FFFFFFFF"
TILE_BG = "FFF4F5F7"
GOOD_TXT, GOOD_BG = "FF0B7A0B", "FFE7F4E7"
BAD_TXT, BAD_BG = "FFA82626", "FFFBE9E9"
WARN_TXT, WARN_BG = "FF8A5A00", "FFFEF3D8"
MUTED = "FF6B7280"
BAR_BLUE = "FF2A78D6"
BORDER = Side(style="thin", color="FFD9DDE3")
THIN_BOX = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

MONEY = "#,##0.00"
INT = "#,##0"
PCT1 = '0.00"%"'

STATUS_HE = {"valid": "תקין", "invalid": "שגוי",
             "no_base": "ללא שכר בסיס פעיל",
             "multi_period": "שתי שורות שכר משולב"}


def _header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, color=NAVY_TEXT, size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 24


def _kpi(ws, row, col, span, title, value, color=None, fmt=None):
    """A merged stat tile: muted title above a large value."""
    c0, c1 = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{c0}{row}:{c1}{row}")
    ws.merge_cells(f"{c0}{row + 1}:{c1}{row + 1}")
    t = ws.cell(row=row, column=col, value=title)
    t.font = Font(size=10, color=MUTED)
    t.alignment = Alignment(horizontal="center", vertical="bottom")
    v = ws.cell(row=row + 1, column=col, value=value)
    v.font = Font(size=16, bold=True, color=color or "FF0B0B0B")
    v.alignment = Alignment(horizontal="center", vertical="top")
    if fmt:
        v.number_format = fmt
    for r in (row, row + 1):
        for cc in range(col, col + span):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TILE_BG)


import re


def _month_from_name(stem):
    """MM/YYYY from an MM.YYYY / YYYY.MM pattern in the file name, else the stem."""
    # Longer alternative first: with (0?[1-9]|1[0-2]) the YYYY.MM form matched
    # the single digit and read "2011.12" as month 1.
    m = re.search(r"(1[0-2]|0?[1-9])[._/](20\d{2})", stem) or \
        re.search(r"(20\d{2})[._/](1[0-2]|0?[1-9])", stem)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        mm, yy = (b, a) if a > 12 else (a, b)
        return f"{mm:02d}/{yy}"
    # Compact MMYY as a standalone token (golmi_1213 -> 12/2013). Not every
    # גולמי carries a תאריך שכר column — the 12/2013 and 1.2008 samples do not
    # — and without this the dashboard labels those files by name, which also
    # sorts them out of chronological order. Guarded: exactly four digits
    # between non-digits, month 01-12, so a bare year like 2024 is rejected.
    m = re.search(r"(?<!\d)(0[1-9]|1[0-2])(\d{2})(?!\d)", stem)
    if m:
        return f"{int(m.group(1)):02d}/{2000 + int(m.group(2))}"
    return stem.split("-", 1)[-1][:12] or stem[:12]


def _sort_month(path):
    """Chronological sort key. Falls back to the name-derived month when the
    file has no תאריך שכר column, so a dateless גולמי sorts by its real month
    instead of landing at the end of the dashboard."""
    d = pay_month_of(path)
    if d:
        return d
    m = re.fullmatch(r"(\d{2})/(\d{4})", _month_from_name(Path(path).stem))
    return datetime(int(m.group(2)), int(m.group(1)), 1) if m else datetime(2099, 1, 1)


def pay_month_of(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    month = None
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        for v in row:
            if isinstance(v, datetime):
                month = v
                break
        if month:
            break
    wb.close()
    return month


def _progim_delta(entry, raw_first, rules):
    """Why the software's verdict differs from a plain current-Progim run —
    one worker's notes, joined. Empty when the two agree exactly."""
    r = entry["result"]
    notes = []
    amt = defaultdict(float)
    for cp in r.components:
        amt[cp.code] += (cp.expected or 0.0)
    raw_vatek = float(raw_first[6] or 0)
    if abs((r.vatek_calculated or 0) - raw_vatek) > 1e-9:
        notes.append(f"ותק קטוע בקובץ: {raw_vatek} תוקן ל-{r.vatek_calculated} "
                     "(ה-Progim מניח ותק מדויק; הקובץ חותך לשתי ספרות)")
    raw_darga = str(raw_first[5] or "").strip()
    if (raw_darga and r.darga_label and raw_darga != str(r.darga_label)
            and engine.normalize_grade_label(raw_darga)
                != engine.normalize_grade_label(r.darga_label)):
        # A bare '+' reposition ('+18'→'18+') is NOT a difference — the Progim
        # resolves both spellings, so we stay silent. Only note a genuine
        # reinterpretation: a קוד-דרגה promoted to its '+' grade by the file's
        # own vote (the labels still differ after normalization).
        notes.append(f"דרגה פורשה כ-'{r.darga_label}' במקום '{raw_darga}' "
                     "(קוד-דרגה קודם לדרגת ה-'+' לפי הצבעת התלושים בקובץ)")
    if r.status == "valid":
        bs = sum(amt[c] for c in engine.BASE_CODES)
        bc = sum(cp.amount for cp in r.components if cp.calculated)
        if bc and abs(bc - bs) > 1.0:
            notes.append(f"הבסיס אושר דרך חלון עיגול הוותק (±0.125 שנה) — "
                         f"חישוב נקודתי כמו ב-Progim היה מראה פער {round(bc - bs, 2)} ₪")
    # NOTE: a component paid at an earlier official phase-in rate is NOT listed
    # here — the Progim resolves the rate by חודש פרישה from its own pulse
    # tables (heskem 2016/2023, tos kibutzi, MANMASH 2010), so a period rate is
    # not a difference between the software and the workbook.
    return "; ".join(notes)


def _gap_reason(code, rules):
    """Short Hebrew 'why' for a gap on a pay code — for the per-code breakdown."""
    if code == "בסיס":
        return "שכר הבסיס אינו תואם — שינוי ותק/דרגה או חודש חלקי"
    if code in (667, 897):
        return "חריגה מהערך התקני לקבוצת הדרגה (חשד להפרשי רטרו)"
    t = (rules.get(code) or {}).get("type")
    if t == "shekel":
        return "סכום קבוע מחוץ לטבלת הערכים הרשמית"
    if t == "max22":
        return "חריגה מנוסחת 4550 (הגבוה מבין 22% למינימום)"
    if t == "percent":
        return "האחוז או בסיס-החישוב אינם תואמים את החוקה"
    if code in (5402, 5524):
        return "סכום שקלי החורג מנורמת קבוצת הדרגה בקובץ"
    return "סטייה מהערך התקני של הרכיב"


# A חוקה amount splits by whether the workbook's figure moves with the
# חודש-פרישה table. Keys are what classify_hukka_amounts.py writes onto
# each rule; an unrecognised/missing value falls through to the "לא נקבע"
# label so an unsettled code is visible instead of silently "fixed".
HUKKA_KIND = {"fixed": "סכום קבוע לכל התקופה",
              "varies": "סכום משתנה מעת לעת",
              # The workbook HAS the figure but picks it by a group the גולמי
              # does not identify (role, tariff, seniority band). Nothing can
              # validate it, so the amount is taken from the משרד האוצר file —
              # kept separate from the amounts a single number does check.
              "group": "תוספת סכומית משתנה לפי בחירת קבוצה"}


def collect(paths, pure=False):
    """Run the engine per file; return (summary, per_emp, code_gaps, recs,
    uncovered).

    pure=True runs the Progim workbook literally (no add-ons) and returns, in
    `recs`, the aggregated recommendations of what to add to the Progim — every
    gap the add-ons would have cleared, itemized, never silently patched.

    `uncovered` lists every pay code found on the slips that the Progim cannot
    COMPUTE — either completely unknown to the workbook, or referenced only as
    an input to other formulas. These are coverage gaps in the product itself
    and get their own sheet in the export.
    """
    lookups = engine.get_lookups()
    rules = engine.get_rules()
    code_gaps = {}   # code -> {"name", "count", "sum"}
    rec_acc = {}     # key -> aggregated recommendation across files
    computable, referenced_only, reported_codes = engine.progim_coverage(rules)
    uncov = {}       # code -> {"name", "rows", "sum", "ministries", "known"}
    seen_codes = {}  # code -> name, every code that appeared on a slip
    files = sorted(paths, key=lambda p: (_sort_month(p), Path(p).name))
    summary, per_emp = [], []
    for path in files:
        d = pay_month_of(path)
        month = d.strftime("%m/%Y") if d else _month_from_name(Path(path).stem)
        short = Path(path).stem.split("-", 1)[-1][:20] or Path(path).stem[:20]
        workers = engine.load_golmi(path)
        entries = engine.run_engine_full(workers, lookups, pure=pure)
        if pure:
            smart = engine.run_engine_full(workers, lookups, pure=False)
            for r in engine.build_progim_recommendations(entries, smart, rules):
                key = (r["category"], r["code"])
                a = rec_acc.get(key)
                if a is None:
                    rec_acc[key] = dict(r)
                else:
                    a["count"] += r["count"]; a["sum"] = round(a["sum"] + r["sum"], 2)
        # Progim coverage gaps: every paid code the workbook cannot compute.
        for e in entries:
            r = e["result"]
            for cp in r.components:
                if cp.code is None:
                    continue
                # Running index of every code seen, for the classification sheet.
                _c = int(cp.code)
                seen_codes[_c] = seen_codes.get(_c) or (str(cp.name) if cp.name else "")
                if _c in computable:
                    continue
                if int(cp.code) in engine.NON_PENSIONABLE:
                    continue   # outside the Progim's scope — not a gap
                code = int(cp.code)
                u = uncov.setdefault(code, {
                    "name": "", "rows": 0, "sum": 0.0, "ministries": set(),
                    "known": code in referenced_only,
                    "reported": code in reported_codes})
                u["rows"] += 1
                u["sum"] += abs(cp.expected or 0.0)
                if r.ministry_name:
                    u["ministries"].add(str(r.ministry_name))
                if cp.name:
                    u["name"] = str(cp.name)
        c = Counter(e["result"].status for e in entries)
        active = c["valid"] + c["invalid"]
        file_start = len(per_emp)   # breakdown is computed after the rows below
        summary.append({
            "month": month, "file": short, "workers": len(workers),
            "valid": c["valid"], "invalid": c["invalid"],
            "no_base": c["no_base"], "multi": c["multi_period"],
            "acc": round(c["valid"] / active * 100, 2) if active else 0.0,
        })
        print(f"  {month}  {len(workers):>7,} עובדים · שגויים {c['invalid']:>5,} "
              f"· {summary[-1]['acc']:.2f}%  ({short})")
        raw_firsts = [rows[0] for rows in workers.values()]
        for e, raw_first in zip(entries, raw_firsts):
            r, flags = e["result"], e["comp_flags"]
            bs = sum((cp.expected or 0.0) for cp in r.components
                     if cp.code in engine.BASE_CODES)
            bc = sum(cp.amount for cp in r.components if cp.calculated)
            # gap in גמולי השתלמות (slip − standard), split so A (667) and B
            # (897) sit next to the base gap; empty when the גמול matched.
            gmul_a = (round(flags[667]["slip"] - flags[667]["expected"], 2)
                      if 667 in flags else None)
            gmul_b = (round(flags[897]["slip"] - flags[897]["expected"], 2)
                      if 897 in flags else None)
            gap_798 = (round(flags[798]["slip"] - flags[798]["expected"], 2)
                       if 798 in flags else None)
            gap_4983 = (round(flags[4983]["slip"] - flags[4983]["expected"], 2)
                        if 4983 in flags else None)
            # 4624 is read from the RAW checks, not the trusted flags: the rule
            # often sits just under the self-calibration gate, so the gap would
            # otherwise be invisible. The column always shows it; the
            # neutralization bucket below still keys off the trusted flags.
            _c4624 = e["comp_checks"].get(4624)
            gap_4624 = (round(_c4624["slip"] - _c4624["expected"], 2)
                        if _c4624 and not _c4624["ok"] else None)
            # result.total/total_diff already include every flagged component's
            # correction (folded in by the engine) — use them as-is.
            total_calc = round(r.total, 2)
            total_diff = round(r.total_diff or 0.0, 2)
            base_diff = round(bc - bs, 2) if bc else None
            # ותק סטודנט (5527): the seniority is computed on the student track,
            # so the regular-vetek recomputation is expectedly off — a known
            # cause, neutralized first (before the base bucket it would land in).
            has_student = any(cp.code == 5527 and (cp.amount or 0) != 0
                              for cp in r.components)
            # ניכוי 6% א"ע — see the bucket below.
            has_1711 = any(cp.code == 1711 and (cp.amount or 0) != 0
                           for cp in r.components)
            # משכ. בסיסית (4140): a basic-salary component with NO formula in
            # the workbook — the most persistent item in the coverage gap
            # (eight files). A slip carrying it has a base the חוקה cannot
            # reproduce, so the recompute is expectedly off. Placed FIRST in
            # the chain at the user's request. Measured before adding: across
            # eight files it catches 5 invalid workers — 4 currently in the
            # base bucket, 1 in 1999, none in "real" — so the true-error count
            # does not move; it only re-attributes.
            has_4140 = any(cp.code == 4140 and (cp.amount or 0) != 0
                           for cp in r.components)
            # ותק קטוע: the file's 2-decimal ותק is not on the quarter grid, so
            # the base recompute rests on a restored (+0.005) value — when the
            # slip still mismatches, the truncated source data is the known
            # cause, not a real pay error.
            vatek_trunc = abs((r.vatek_calculated or 0)
                              - float(raw_first[6] or 0)) > 1e-9
            # Error category, deduplicated by priority: a worker with several
            # gaps counts once, under the first bucket in this order — so
            # neutralizing each known cause leaves only the true unknowns.
            err_cat = None
            if r.status == "invalid":
                if has_4140:
                    err_cat = "b4140"
                elif has_student:
                    err_cat = "student"
                elif vatek_trunc:
                    err_cat = "vatek"
                elif base_diff is not None and abs(base_diff) > 1.0:
                    err_cat = "base"
                elif 667 in flags or 897 in flags:
                    err_cat = "gmul"
                elif has_1711:
                    # ניכוי 6% א"ע (1711): a DEDUCTION, not a payment — its
                    # amount is negative on the slip and the Progim, which
                    # covers pensionable pay, has no formula for it. A worker
                    # carrying it is neutralized rather than counted as an
                    # error. Placed where the user asked, right after גמול.
                    err_cat = "d1711"
                elif gap_4624 is not None:
                    # הסכם 1999: הפערים נובעים מהפרשי רטרו (בולט אצל עובדים
                    # בעלי ותק נמוך) — סיבה ידועה, מנוטרלת. נבדק מול הבדיקה
                    # הגולמית ולא מול הדגלים, כי הכלל יושב לרוב מתחת לסף
                    # הכיול העצמי ואחרת הנטרול לא היה נכנס לפעולה כלל.
                    # ממוקם מיד אחרי גמול השתלמות ולפני דריכות/גמול מנהל:
                    # 4624 יושב בבסיס החישוב של 798 ושל שאר האחוזיות, ולכן
                    # הפרש רטרו ב-1999 מתגלגל אליהן — הייחוס הוא לשורש ולא
                    # לתסמין.
                    err_cat = "h1999"
                elif 798 in flags:
                    err_cat = "brich"
                elif 4983 in flags:
                    err_cat = "mnhal"
                elif 741 in flags:
                    # בוררות מיסים: תקין ב-Progim; פערים בקבצים נובעים מחישובי
                    # הפרשים (רטרו) — סיבה ידועה, מנוטרלת.
                    err_cat = "borerut"
                elif 705 in flags:
                    # מקצועית מיסים (705): סכום קבוע בחוקה (158.63). הפערים
                    # בקבצים ההיסטוריים הם כפולות שלמות של הסכום — 12×, 4×, 3× —
                    # כלומר תשלומי רטרו שנדחסו לשורה אחת, לא שגיאות. הגולמי
                    # אינו נושא דגל רטרו ולכן אי אפשר להפריד אותם בקוד.
                    # ממוקם מיד אחרי בוררות מיסים לפי בקשת המשתמש.
                    err_cat = "mikzoit"
                elif 738 in flags:
                    # תוספת אחוז יום (738): הרכיב הוגדר במנוע רק ב-5.8.2026 —
                    # הוא היה מוגדר בחוברת (tosafot!BZ) אך המחלץ פספס אותו.
                    # ממוקם מיד אחרי מקצועית מיסים לפי בקשת המשתמש.
                    # נמדד לפני ההוספה: 26 עובדים / ₪1,132 בארבעה קבצים,
                    # וכל אחד מהם נושא את 738 לבדו — הדלי אינו בולע דבר.
                    err_cat = "ahuz_yom"
                elif 600 in flags:
                    # תוספת בית חולים (600): סכום קבוע בחוקה (75.26); הסוטים
                    # ממנו הם ברובם פעימה קודמת של הסכום, שהחוקה אינה שומרת.
                    err_cat = "bhol"
                elif 5340 in flags:
                    # בית חולים מאוחדת (5340): סכום קבוע 480.1 בחוקה. הרכיב
                    # הומר מ-`reported` ל-`shekel` ב-5.8.2026 ולכן נבדק לראשונה;
                    # ממוקם מיד אחרי תוספת בית חולים לפי בקשת המשתמש.
                    # ‼ נמדד לפני ההוספה: הדלי קולט 64 עובדים ו-₪18,424 בחמישה
                    # קבצים, ומתוכם 27 נושאים סמל כושל נוסף — 626 (15), 728 (8),
                    # 959 (2), 736 (1), 5216 (1). כלומר הוא **בולע** שגיאות של
                    # רכיבים אחרים, בניגוד לדלי 4140. ראו PROGIM_IMPROVEMENTS.md.
                    err_cat = "bmeuhedet"
                elif 875 in flags:
                    # תוספת בתי משפט (875): ממוקם מיד אחרי תוספת בית חולים לפי
                    # בקשת המשתמש. הכלל רשום ב-PROGIM_FIXES.md §4 כ"כמעט-מתאמת"
                    # (2,819 נושאים, 90 לא-תואמים ב-0108). נמדד לפני ההוספה:
                    # הדלי קולט 25 עובדים בסך כל ששת הקבצים, כולם ב-12/2012,
                    # וכל אחד מהם נושא את 875 לבדו — שום סמל אחר אינו נבלע איתו.
                    err_cat = "bmish"
                elif 5524 in flags:
                    # תוספת שקלית 2023 (5524): ממוקם מיד אחרי תוספת בית משפט
                    # לפי בקשת המשתמש. הרכיב עצמו תקין — ‎14,623‎ מתוך ‎14,816‎
                    # עובדי משרה מלאה ב-12/2023 משלמים בדיוק את ‎₪400‎ שבחוקה.
                    # הסטיות אינן תעריף שגוי אלא **יחסיות ימים**: מתוך ‎279‎
                    # פערים, ‎77‎ שליליים (היפוכי רטרו) ו-‎131‎ כפולות מדויקות
                    # של ‎₪400/30 = ₪13.33‎. החוברת מגדירה את ‎5524‎ כסכום קבוע
                    # בלי כלל לחודש חלקי, ולכן כל תלוש חלקי נראה כסטייה.
                    # ראו PROGIM_FIXES.md §19 — מה שיבטל את הדלי הזה הוא
                    # הצהרת יחסיות ‎1/30‎ ליום בחוברת.
                    err_cat = "shk2023"
                else:
                    err_cat = "real"
                # Per-code gap tally for the exec dashboard: base (when off) plus
                # every flagged component, each with its ₪ magnitude.
                def _bump(code, name, amt):
                    g = code_gaps.setdefault(code, {"name": name, "count": 0, "sum": 0.0})
                    g["count"] += 1
                    g["sum"] += abs(amt or 0.0)
                    if name:
                        g["name"] = name
                if base_diff is not None and abs(base_diff) > 1.0:
                    _bump("בסיס", "שכר בסיס", base_diff)
                for k, v in flags.items():
                    _bump(k, v["name"], v["slip"] - v["expected"])
            per_emp.append({
                "month": month, "file": short, "worker_id": r.worker_id,
                "ministry": r.ministry_name, "darga": r.darga_label,
                "vatek": r.vatek_calculated, "job_pct": r.job_pct,
                "full_time": (r.job_pct or 1.0) >= 0.999,
                "err_cat": err_cat, "neutral_he": NEUTRAL_HE.get(err_cat, ""),
                "base_slip": round(bs, 2),
                "base_calc": round(bc, 2) if bc else None,
                "base_diff": base_diff,
                "gmul_a": gmul_a, "gmul_b": gmul_b, "gap_798": gap_798,
                "gap_4983": gap_4983, "gap_4624": gap_4624,
                "total_slip": r.expected_total, "total_calc": total_calc,
                "total_diff": total_diff, "status": r.status,
                # Direction drives the process: a slip paid BELOW the rulebook
                # is money owed to the worker (retro payment); paid ABOVE is a
                # recovery case. Materiality separates actionable gaps from
                # agora-level noise nobody will open a file for.
                "direction": ("שולם בחסר" if total_diff > 0.5 else
                              "שולם ביתר" if total_diff < -0.5 else ""),
                "material": ("כן" if abs(total_diff) >= MATERIALITY else "לא"),
                "flags": "; ".join(
                    f"{k} ({v['name']}): {v['slip']} במקום {v['expected']}"
                    for k, v in sorted(flags.items())),
                "diag": "; ".join(r.errors),
                "progim_delta": _progim_delta(e, raw_first, rules),
            })
        # Full-time breakdown for the dashboard: part-timers are neutralized,
        # and each invalid full-timer lands in exactly ONE bucket (base >
        # gmul > דריכות > אמיתי) so the buckets never double-count.
        rows_f = per_emp[file_start:]
        ft = [x for x in rows_f if x["full_time"]]
        s = summary[-1]
        s["part_time"] = len(rows_f) - len(ft)
        s["ft"] = len(ft)
        s["ft_valid"] = sum(1 for x in ft if x["status"] == "valid")
        s["ft_no_base"] = sum(1 for x in ft if x["status"] == "no_base")
        s["ft_multi"] = sum(1 for x in ft if x["status"] == "multi_period")
        # The dashboard buckets cover the WHOLE file, not just full-timers.
        # Counting them on full-timers alone left every non-valid part-timer
        # (102 invalid + 7 retro on the 0108 file) with no column at all: they
        # vanished into "משרה חלקית" and the partition only closed because that
        # column swallowed them. Now: no-base + retro + the buckets + valid =
        # the worker total, and משרה חלקית is a descriptive "of which" column
        # standing OUTSIDE the partition.
        for cat, key in (("b4140", "inv_b4140"),
                         ("student", "inv_student"), ("vatek", "inv_vatek"),
                         ("base", "inv_base"), ("gmul", "inv_gmul"),
                         ("d1711", "inv_d1711"),
                         ("h1999", "inv_h1999"), ("brich", "inv_brich"),
                         ("mnhal", "inv_mnhal"), ("borerut", "inv_borerut"),
                         ("mikzoit", "inv_mikzoit"),
                         ("ahuz_yom", "inv_ahuz_yom"),
                         ("bhol", "inv_bhol"), ("bmeuhedet", "inv_bmeuhedet"),
                         ("bmish", "inv_bmish"),
                         ("shk2023", "inv_shk2023"),
                         ("real", "inv_real")):
            s[key] = sum(1 for x in rows_f
                         if x["status"] == "invalid" and x["err_cat"] == cat)
        # % is of the WHOLE file, and now every worker in that denominator has
        # actually been bucketed — reads as "x% of all slips are a real error".
        s["real_pct"] = (round(s["inv_real"] / s["workers"] * 100, 2)
                         if s["workers"] else 0.0)
    code_gap_list = sorted(
        ({"code": code, "name": g["name"], "count": g["count"],
          "sum": round(g["sum"], 2), "reason": _gap_reason(code, rules)}
         for code, g in code_gaps.items()),
        key=lambda x: -x["count"])
    recs = sorted(rec_acc.values(), key=lambda r: -r["count"])
    # ---- classification of every known code, by the three kinds + out-of-scope
    codes_index = []
    for code in sorted(set(seen_codes) | {int(k) for k in rules}):
        rule = rules.get(code)
        name = seen_codes.get(code) or (rule or {}).get("name", "") or ""
        if code in engine.NON_PENSIONABLE:
            kind, note = "לא משתתף בחישובים", "מחוץ לתחולת ה-Progim — רכיב שאינו פנסיוני (תקין)"
        elif code in (engine.CODE_YESOD, engine.CODE_VETEK_TOSEFET,
                      engine.CODE_COMBINED_BASE):
            kind, note = "מחושב לפי נוסחה", "שכר יסוד × מקדם ותק × חלקיות"
        elif code in set(engine.GMUL_A_CODES) | set(engine.GMUL_B_CODES):
            kind, note = "מחושב לפי נוסחה", "גמול השתלמות — ערך תקני לקבוצת הדרגה"
        elif rule is None:
            kind, note = "לא מוגדר בחוברת", "פנסיוני אך חסר בחוברת — פער לסגירה"
        elif rule.get("origin") == "manual":
            kind, note = "סכום מוזן ידנית", rule.get("source", "")
        elif rule.get("origin") == "hukka":
            # A חוקה amount is one of two things, and the difference decides
            # whether one figure can validate every month: either the workbook
            # holds a single figure for the whole period, or it holds a
            # חודש-פרישה table whose figure has actually moved. Read out of the
            # workbook by tools/classify_hukka_amounts.py — never assumed; a
            # code the workbook cannot settle stays in its own third label
            # rather than being guessed into one of the two.
            kind = HUKKA_KIND.get(rule.get("amount_period"),
                                  "סכום לפי חוקה — לא נקבע")
            parts = []
            if rule.get("type") == "shekel":
                parts.append(f"סכום בחוקה: {rule.get('amounts')}")
            else:
                parts.append("סכום בחוקה — לא נפתר אוטומטית, מתקבל כמות-שהוא")
            if rule.get("amount_period_note"):
                parts.append(rule["amount_period_note"])
            note = " · ".join(parts)
        elif rule.get("rate_group"):
            # A percentage whose RATE is picked by a group the גולמי does not
            # carry (4406: a 1/2/3 code hand-entered in Netunei Gimlai). The
            # base is computable, so unlike the shekel case the amount can
            # still be tested against the admissible rates — but which rate is
            # correct for a given worker cannot be decided from the file.
            kind = "תוספת אחוזית משתנה לפי בחירת קבוצה"
            note = ("שער " + ", ".join(f"{x*100:g}%" for x in rule.get("rates", []))
                    + " · " + rule["rate_group"])
        else:
            kind = "מחושב לפי נוסחה"
            if rule.get("type") == "percent":
                note = "שער " + ", ".join(f"{x*100:g}%" for x in rule.get("rates", []))
            elif rule.get("type") == "max22":
                note = "MAX(22% מהמשולב+99, רצפת המשרד)"
            elif rule.get("type") == "minimum":
                note = "השלמה לשכר מינימום"
            else:
                note = ""
        codes_index.append({"code": code, "name": name, "kind": kind, "note": note})

    uncovered = sorted(
        ({"code": code, "name": u["name"], "rows": u["rows"],
          "sum": round(u["sum"], 2), "known": u["known"],
          "reported": u["reported"],
          "ministries": ", ".join(sorted(u["ministries"])[:3])
                        + ("…" if len(u["ministries"]) > 3 else "")}
         for code, u in uncov.items()),
        key=lambda x: -x["sum"])
    return summary, per_emp, code_gap_list, recs, uncovered, codes_index


# (key, header, width, number-format, is-gap-cell). Gap cells are red-tinted
# when they carry a real value so the eye lands on them.
# Hebrew label per neutralization bucket, shown in the work queue so a
# neutralized row is visible and filterable instead of silently dropped.
# The neutralization chain in display order — must match the err_cat priority
# in collect() and the dashboard column order. Stated on the report itself so
# a reader can see which bucket claimed a worker first.
CHAIN_HE = ('ללא בסיס ← שתי שורות שכר משולב ← משכ. בסיסית (4140) ← ותק סטודנט ← ותק קטוע ← בסיס '
            '← גמול ← ניכוי 6% א"ע ← תוספת 1999 ← דריכות ← גמול מנהל '
            '← בוררות מיסים ← מקצועית מיסים ← תוספת אחוז יום ← תוספת בית חולים '
            '← בית חולים מאוחדת ← תוספת בית משפט ← תוספת שקלית 2023 ← שגיאה אמיתית ← תקין')

NEUTRAL_HE = {"b4140": "משכ. בסיסית 4140", "student": "ותק סטודנט", "vatek": "ותק קטוע", "base": "שכר בסיס",
              "gmul": "גמול השתלמות", "d1711": "ניכוי 6% א\"ע",
              "brich": "דריכות בי\"ח",
              "mnhal": "גמול מנהל", "borerut": "בוררות מיסים",
              "mikzoit": "מקצועית מיסים",
              "bhol": "תוספת בית חולים",
              "ahuz_yom": "תוספת אחוז יום",
              "bmeuhedet": "בית חולים מאוחדת",
              "bmish": "תוספת בית משפט",
              "shk2023": "תוספת שקלית 2023",
              "h1999": "תוספת 1999", "real": ""}

# A gap below this is real but not worth opening a case for — the report keeps
# it, and marks it so the work queue can be filtered to what matters.
MATERIALITY = 100.0

EMP_COLS = [
    ("month", "חודש שכר", 11, None, False), ("file", "קובץ", 18, None, False),
    ("worker_id", "מסד עובד", 12, INT, False), ("ministry", "משרד", 22, None, False),
    ("darga", "דרגה", 8, None, False), ("vatek", "ותק", 8, None, False),
    ("job_pct", "חלקיות", 8, None, False), ("base_slip", "בסיס בתלוש", 13, MONEY, False),
    ("base_calc", "בסיס מחושב", 13, MONEY, False), ("base_diff", "הפרש בסיס", 12, MONEY, True),
    ("gmul_a", "פער גמול א'", 12, MONEY, True), ("gmul_b", "פער גמול ב'", 12, MONEY, True),
    ("gap_798", "פער דריכות בי\"ח", 13, MONEY, True),
    ("gap_4983", "פער גמול מנהל", 12, MONEY, True),
    ("gap_4624", "פער תוספת 1999", 13, MONEY, True),
    ("total_slip", "סכום בתלוש", 13, MONEY, False), ("total_calc", "סכום מחושב", 13, MONEY, False),
    ("total_diff", "הפרש כולל", 12, MONEY, True),
    ("direction", "כיוון", 11, None, False), ("material", "מהותי", 8, None, False),
    ("neutral_he", "סיבת נטרול", 14, None, False),
    ("status_he", "סטטוס", 16, None, False),
    ("flags", "רכיבים חריגים", 30, None, False), ("diag", "אבחון", 30, None, False),
    ("progim_delta", "שוני מול Progim — והסבר", 42, None, False),
]


def _emp_sheet(wb, title, rows, table_name, highlight_invalid):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "D2"            # keep month/file/מסד visible when scrolling ₪ cols
    # The work queue carries a cumulative-₪ column; the full roster does not.
    cols = list(EMP_COLS)
    if rows and "cum_pct" in rows[0]:
        cols.append(("cum_pct", "% מצטבר מהחשיפה", 15, "0.0%", False))
    _header_row(ws, 1, [he for _, he, _, _, _ in cols],
                [w for _, _, w, _, _ in cols])
    inv_font = Font(color=BAD_TXT, bold=True)
    inv_fill = PatternFill("solid", fgColor=BAD_BG)
    ok_font = Font(color=GOOD_TXT)
    warn_font = Font(color=WARN_TXT)
    for r_i, row in enumerate(rows, start=2):
        vals = [row.get(k) if k != "status_he" else STATUS_HE[row["status"]]
                for k, _, _, _, _ in cols]
        for c_i, ((key, _, _, fmt, is_gap), v) in enumerate(zip(cols, vals), start=1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            if fmt:
                cell.number_format = fmt
            if key == "status_he":
                st = row["status"]
                cell.font = (inv_font if st == "invalid"
                             else ok_font if st == "valid" else warn_font)
                if st == "invalid":
                    cell.fill = inv_fill
            elif is_gap and isinstance(v, (int, float)) and abs(v) > 1:
                # any real gap gets a red tint — the eye lands on it directly
                cell.font = inv_font
                cell.fill = inv_fill
        if highlight_invalid and row["status"] == "invalid":
            ws.cell(row=r_i, column=1).fill = inv_fill
    if rows:
        ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight15",
                                              showRowStripes=True)
        ws.add_table(table)
    return ws


def _anomaly(row):
    """₪ magnitude of a row's gap: larger of the base gap and the corrected-total
    gap (the latter folds in component corrections incl. גמול), so a component-
    only mismatch still ranks."""
    return max(abs(row["base_diff"] or 0.0), abs(row["total_diff"] or 0.0))


def _month_key(m):
    try:
        mm, yy = str(m).split("/")
        return int(yy) * 12 + int(mm)
    except ValueError:
        return 0


def compute_flips(per_emp):
    """Workers whose slip flipped valid→invalid between consecutive months."""
    hist = defaultdict(list)
    for r in per_emp:
        hist[r["worker_id"]].append(r)
    flips = []
    for rows in hist.values():
        if len(rows) < 2:
            continue
        rows.sort(key=lambda r: _month_key(r["month"]))
        for prev, cur in zip(rows, rows[1:]):
            if (prev["status"] == "valid" and cur["status"] == "invalid"
                    and _month_key(prev["month"]) != _month_key(cur["month"])):
                flips.append({"id": cur["worker_id"], "ministry": cur["ministry"],
                              "darga": cur["darga"], "from": prev["month"],
                              "to": cur["month"], "diff": _anomaly(cur)})
    return sorted(flips, key=lambda f: -f["diff"])


def write_workbook(summary, per_emp, out_path, code_gaps=None, recs=None,
                   uncovered=None, codes_index=None, skip_per_employee=False):
    wb = openpyxl.Workbook()
    anom_by_file = defaultdict(float)
    anom_total = 0.0
    for r in per_emp:
        if r["status"] == "invalid":
            v = _anomaly(r)
            anom_by_file[(r["month"], r["file"])] += v
            anom_total += v

    # ---- לוח בקרה ------------------------------------------------------------
    ws = wb.active
    ws.title = "לוח בקרה"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:Q1")
    t = ws["A1"]
    t.value = "בדיקת התאמת תלושים — דוח מאוחד"
    t.font = Font(size=16, bold=True, color=NAVY)
    ws.merge_cells("A2:Q2")
    s = ws["A2"]
    s.value = (f"{len(summary)} קבצים · הופק "
               f"{datetime.now().strftime('%d/%m/%Y %H:%M')} · "
               "כל עובד נספר בעמודה אחת בדיוק: " + CHAIN_HE + ". סכום "
               "העמודות (כולל 'תקין' בסוף, למעט 'משרה חלקית') = "
               "סה\"כ העובדים. עמודת 'משרה חלקית (מתוכם)' היא תיאורית "
               "בלבד — אותם עובדים נספרים גם בעמודת הסיבה שלהם. "
               "% שגויים אמיתיים = אמיתיים חלקי סה\"כ העובדים.")
    s.font = Font(size=10, color=MUTED)
    # Coverage warning on the dashboard itself — the Progim is the product, so a
    # hole in it must be visible on page one, not only in its own sheet.
    if uncovered:
        _rep = [u for u in uncovered if u.get("reported")]
        _gap = [u for u in uncovered if not u.get("reported")]
        s_gap = round(sum(u["sum"] for u in _gap))
        s_rep = round(sum(u["sum"] for u in _rep))
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=23)
        w = ws.cell(row=3, column=1,
                    value=(f"\u26a0 {len(_gap)} סמלים (₪{s_gap:,}) ללא נוסחה ב-Progim "
                           f"ואינם מוצהרים כמוזנים — חור בכיסוי החוברת. "
                           f"בנוסף {len(_rep)} סמלים (₪{s_rep:,}) מוזנים מהקובץ "
                           f"כפי שה-Progim מגדיר (תקין). ראה גיליון \"חסר ב-Progim\""))
        w.font = Font(bold=True, size=10,
                      color=BAD_TXT if _gap else WARN_TXT)
        w.fill = PatternFill("solid", fgColor=BAD_BG if _gap else WARN_BG)

    tot = Counter()
    for r in summary:
        for k in ("workers", "valid", "invalid", "no_base", "multi"):
            tot[k] += r[k]
    active = tot["valid"] + tot["invalid"]
    acc = round(tot["valid"] / active * 100, 2) if active else 0.0
    _kpi(ws, 4, 1, 1, "סה\"כ עובדים", tot["workers"], fmt=INT)
    _kpi(ws, 4, 2, 1, "תקינים", tot["valid"], GOOD_TXT, INT)
    for k in ("part_time", "ft", "ft_valid", "ft_no_base", "ft_multi",
              "inv_b4140", "inv_student", "inv_vatek", "inv_base", "inv_gmul", "inv_d1711",
              "inv_h1999",
              "inv_brich", "inv_mnhal", "inv_borerut", "inv_mikzoit",
              "inv_ahuz_yom", "inv_bhol", "inv_bmeuhedet", "inv_bmish",
              "inv_real"):
        tot[k] = sum(r.get(k, 0) for r in summary)
    _real_pct = (tot["inv_real"] / tot["workers"]) if tot["workers"] else 0.0
    _kpi(ws, 4, 3, 1, "שגויים אמיתיים", tot["inv_real"], BAD_TXT, INT)
    _kpi(ws, 4, 4, 1, "% שגויים אמיתיים", _real_pct,
         GOOD_TXT if _real_pct <= 0.01 else BAD_TXT, "0.00%")
    # Direction split — two different processes: money owed to workers vs money
    # to recover. Gross exposure, not net, so neither side is hidden.
    _under = sum(r["total_diff"] for r in per_emp
                 if r["err_cat"] == "real" and (r["total_diff"] or 0) > 0.5)
    _over = sum(-r["total_diff"] for r in per_emp
                if r["err_cat"] == "real" and (r["total_diff"] or 0) < -0.5)
    _kpi(ws, 4, 5, 1, "שולם בחסר ₪ (לתשלום)", round(_under), WARN_TXT, INT)
    _kpi(ws, 4, 6, 1, "שולם ביתר ₪ (להשבה)", round(_over), BAD_TXT, INT)
    _kpi(ws, 4, 7, 1, "חשיפה ברוטו ₪", round(_under + _over), BAD_TXT, INT)

    head_r = 7
    # Neutralization chain (each invalid full-timer counts once, in this order):
    # ותק סטודנט → ותק קטוע → בסיס → גמול → תוספת 1999 → דריכות → גמול מנהל →
    # בוררות מיסים. What remains is the REAL error count. 1999 sits directly
    # after גמול because 4624 is part of the base of 798 and the other percent
    # tosafot, so a retro difference there propagates into them — attribute to
    # the root, not the symptom.
    # רטרו/רב-תקופתי is a SEPARATE class (base code appears twice — the slip
    # merges two pay periods), not part of the chain, shown for reference only.
    # A clean PARTITION: columns D..P are mutually exclusive and sum to C
    # (עובדים) — presentable without reconciliation notes.
    # "תקין" sits LAST, after the %, so the problem columns read as one block.
    # The partition is unchanged — it is simply no longer contiguous: the
    # categories are D..O plus Q.
    labels = ["חודש שכר", "קובץ", "עובדים", "משרה חלקית (מתוכם)", "ללא בסיס",
              "שתי שורות שכר משולב", "שגויי משכ. בסיסית 4140",
              "שגויי ותק סטודנט", "שגויי ותק קטוע",
              "שגויי בסיס", "שגויי גמול", "שגויי ניכוי 6% א\"ע",
              "שגויי תוספת 1999", "שגויי דריכות",
              "שגויי גמול מנהל", "שגויי בוררות מיסים",
              "שגויי מקצועית מיסים", "שגויי תוספת אחוז יום",
              "שגויי תוספת בית חולים",
              "שגויי בית חולים מאוחדת", "שגויי תוספת בית משפט",
              "שגויי תוספת שקלית 2023",
              "שגויים אמיתיים", "% שגויים אמיתיים", "תקין"]
    _header_row(ws, head_r, labels,
                [11, 18, 10, 10, 10, 13, 14, 11, 11, 10, 10, 13, 11, 10, 11, 11, 13, 14, 13, 14, 13, 15, 12, 13, 11])
    for i, r in enumerate(summary, start=head_r + 1):
        vals = [r["month"], r["file"], r["workers"], r.get("part_time", 0),
                r.get("no_base", 0), r.get("multi", 0),
                r.get("inv_b4140", 0), r.get("inv_student", 0), r.get("inv_vatek", 0),
                r.get("inv_base", 0), r.get("inv_gmul", 0),
                r.get("inv_d1711", 0), r.get("inv_h1999", 0),
                r.get("inv_brich", 0), r.get("inv_mnhal", 0),
                r.get("inv_borerut", 0), r.get("inv_mikzoit", 0),
                r.get("inv_ahuz_yom", 0), r.get("inv_bhol", 0),
                r.get("inv_bmeuhedet", 0),
                r.get("inv_bmish", 0),
                r.get("inv_shk2023", 0),
                r.get("inv_real", 0),
                r.get("real_pct", 0.0) / 100, r.get("valid", 0)]
        for c_i, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c_i, value=v)
            cell.border = THIN_BOX
            if 3 <= c_i <= 23 or c_i == 25:
                cell.number_format = INT
            if c_i == 25:
                cell.font = Font(color=GOOD_TXT)
            if c_i in (7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22) and v:
                cell.font = Font(color=WARN_TXT)
            if c_i == 23 and v:
                cell.font = Font(color=BAD_TXT, bold=True)
            if c_i == 24:
                cell.number_format = "0.00%"
    last = head_r + len(summary)
    trow = last + 1
    real_pct_tot = _real_pct
    # Same order as `labels`: תקין is the LAST column, % second to last.
    tvals = ["סה\"כ", "", tot["workers"], tot["part_time"], tot["no_base"],
             tot["multi"], tot["inv_b4140"], tot["inv_student"], tot["inv_vatek"],
             tot["inv_base"], tot["inv_gmul"], tot["inv_d1711"],
             tot["inv_h1999"],
             tot["inv_brich"], tot["inv_mnhal"], tot["inv_borerut"],
             tot["inv_mikzoit"], tot["inv_ahuz_yom"], tot["inv_bhol"],
             tot["inv_bmeuhedet"],
             tot["inv_bmish"],
             tot["inv_shk2023"],
             tot["inv_real"],
             real_pct_tot, tot["valid"]]
    for c_i, v in enumerate(tvals, start=1):
        cell = ws.cell(row=trow, column=c_i, value=v)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="double", color=NAVY))
        if 3 <= c_i <= 23 or c_i == 25:
            cell.number_format = INT
        if c_i == 24:
            cell.number_format = "0.00%"
        if c_i == 23:
            cell.font = Font(bold=True, color=BAD_TXT)
    rng = f"X{head_r + 1}:X{last}"    # % שגויים אמיתיים
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThanOrEqual", formula=["0.01"],
        font=Font(color=GOOD_TXT), fill=PatternFill("solid", fgColor=GOOD_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["0.0101", "0.05"],
        font=Font(color=WARN_TXT), fill=PatternFill("solid", fgColor=WARN_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0.05"],
        font=Font(color=BAD_TXT), fill=PatternFill("solid", fgColor=BAD_BG)))
    ws.conditional_formatting.add(
        f"C{head_r + 1}:C{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=BAR_BLUE, showValue=True))
    ws.conditional_formatting.add(
        f"V{head_r + 1}:V{last}",          # שגויים אמיתיים
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="FFD03B3B", showValue=True))

    # ---- פערים לפי סמל שכר (a second table, to the right of the per-file one) ---
    if code_gaps:
        cbase = 25   # column Y (leaves a gap after the 23-column per-file table)
        heads = ["סמל", "שם רכיב", "כמות פערים", "שווי ₪", "הסיבה לפער"]
        widths = [9, 22, 12, 13, 40]
        for j, (h, w) in enumerate(zip(heads, widths)):
            c = ws.cell(row=head_r, column=cbase + j, value=h)
            c.font = Font(bold=True, color=NAVY_TEXT, size=11)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BOX
            ws.column_dimensions[get_column_letter(cbase + j)].width = w
        cl = ws.cell(row=head_r - 1, column=cbase, value="פערים לפי סמל שכר — כמה וסיבה")
        cl.font = Font(bold=True, size=12, color=NAVY)
        for gi, g in enumerate(code_gaps, start=head_r + 1):
            vals = [g["code"], g["name"], g["count"], round(g["sum"]), g["reason"]]
            for j, v in enumerate(vals):
                cell = ws.cell(row=gi, column=cbase + j, value=v)
                cell.border = THIN_BOX
                if j in (2, 3):
                    cell.number_format = INT
                if j == 2 and g["count"]:
                    cell.font = Font(bold=True, color=BAD_TXT)
        gl = head_r + len(code_gaps)
        ws.conditional_formatting.add(
            f"{get_column_letter(cbase + 2)}{head_r + 1}:{get_column_letter(cbase + 2)}{gl}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color="FFD03B3B", showValue=True))

    # ---- שגויים לבדיקה (ממוין לפי ₪) --------------------------------------------
    # Every invalid full-timer stays in the queue — a neutralized row is
    # LABELLED (סיבת נטרול), never dropped, so the sheet remains the full
    # work list. Real errors sort first, then by ₪ within each group.
    inv = [r for r in per_emp if r["status"] == "invalid" and r["full_time"]]
    inv.sort(key=lambda r: (r["err_cat"] != "real", -_anomaly(r)))
    # Cumulative share of the ₪ exposure, so the queue answers "how far down do
    # I have to work to cover most of the money".
    _tot_anom = sum(_anomaly(r) for r in inv) or 1.0
    _run = 0.0
    for r in inv:
        _run += _anomaly(r)
        r["cum_pct"] = round(_run / _tot_anom, 4)
    _emp_sheet(wb, "שגויים לבדיקה", inv, "Invalids", highlight_invalid=False)

    # ---- ⚠ חסר ב-Progim: paid codes the workbook cannot compute -----------------
    # The product being sold is the Progim. Every code here is money the payroll
    # pays that the workbook has NO formula for — the engine accepts it as
    # reported, unchecked. This sheet is the fix-list for the workbook itself.
    if uncovered:
        wsu = wb.create_sheet("חסר ב-Progim", 1)   # right after the dashboard
        wsu.sheet_view.rightToLeft = True
        wsu.sheet_properties.tabColor = "A82626"
        t = wsu.cell(row=1, column=1,
                     value="רכיבים שה-Progim אינו מחשב בנוסחה — "
                           "אפור = מוזן מהקובץ (כך גם ב-Progim, תקין); "
                           "אדום/לבן = חור בכיסוי החוברת")
        t.font = Font(bold=True, size=12, color=BAD_TXT)
        wsu.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        s = wsu.cell(row=2, column=1,
                     value="'מוזכר כקלט בלבד' = הסמל מופיע בכללים שחולצו רק בתוך "
                           "בסיס/קיזוז/ספירת-מינימום של רכיב אחר; "
                           "'לא נמצא בכללים' = הסמל אינו מופיע באף כלל. "
                           "שים לב: שתי התוויות מתארות את הכללים שחולצו, לא את "
                           "החוברת עצמה — ייתכן שהחוברת כן מגדירה את הרכיב "
                           "והחילוץ פשוט לא הגיע אליו (כך היה עם 4180). "
                           "בשני המקרים אין נוסחה שמחשבת אותו כאן — לבדוק בחוברת.")
        s.font = Font(size=9, color=MUTED)
        wsu.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        tot_u = round(sum(u["sum"] for u in uncovered))
        c3 = wsu.cell(row=3, column=1,
                      value=f"{len(uncovered)} סמלים · ₪{tot_u:,} שלא נבדקו")
        c3.font = Font(bold=True, size=10, color=NAVY)
        wsu.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        _header_row(wsu, 4, ["סמל", "שם הרכיב", "סטטוס ב-Progim", "שורות",
                             "₪ (מוחלט)", "גופים"],
                    [8, 20, 18, 9, 13, 40])
        for i, u in enumerate(uncovered, start=5):
            status = ("מוזן מהקובץ — כך ב-Progim" if u.get("reported")
                      else "מוזכר כקלט בלבד" if u["known"]
                      else "לא נמצא בכללים")
            vals = [u["code"], u["name"], status,
                    u["rows"], u["sum"], u["ministries"]]
            for c_i, v in enumerate(vals, start=1):
                cell = wsu.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
                if c_i == 4:
                    cell.number_format = INT
                if c_i == 5:
                    cell.number_format = MONEY
            if u.get("reported"):
                for c_i in range(1, 7):
                    wsu.cell(row=i, column=c_i).fill = \
                        PatternFill("solid", fgColor="FFEFEFEF")
            elif not u["known"]:   # completely unknown — the louder class
                wsu.cell(row=i, column=3).font = Font(bold=True, color=BAD_TXT)
                for c_i in range(1, 7):
                    wsu.cell(row=i, column=c_i).fill = \
                        PatternFill("solid", fgColor=BAD_BG)
        wsu.freeze_panes = "A5"

    # ---- סיווג סמלי שכר: every code, ascending, by the three kinds -------------
    if codes_index:
        wsc = wb.create_sheet("סיווג סמלי שכר", 2)
        wsc.sheet_view.rightToLeft = True
        t = wsc.cell(row=1, column=1,
                     value="כל סמלי השכר לפי סדר עולה — סיווג לפי אופן קביעת הסכום")
        t.font = Font(bold=True, size=12, color=NAVY)
        wsc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        KIND_COLOR = {"מחושב לפי נוסחה": GOOD_BG,
                      "סכום קבוע לכל התקופה": "FFEAF2FB",
                      # amber: a moving figure cannot be checked with one number
                      "סכום משתנה מעת לעת": WARN_BG,
                      "תוספת סכומית משתנה לפי בחירת קבוצה": WARN_BG,
                      "תוספת אחוזית משתנה לפי בחירת קבוצה": WARN_BG,
                      "סכום לפי חוקה — לא נקבע": BAD_BG,
                      "סכום מוזן ידנית": "FFEFEFEF",
                      "לא משתתף בחישובים": "FFF4F5F7",   # מחוץ לתחולה — תקין
                      "לא מוגדר בחוברת": BAD_BG}          # הפער שנסגר בהדרגה
        cnt = Counter(x["kind"] for x in codes_index)
        sub = wsc.cell(row=2, column=1, value=" · ".join(
            f"{k}: {cnt[k]}" for k in ("מחושב לפי נוסחה",
                                       "סכום קבוע לכל התקופה",
                                       "סכום משתנה מעת לעת",
                                       "תוספת סכומית משתנה לפי בחירת קבוצה",
                                       "תוספת אחוזית משתנה לפי בחירת קבוצה",
                                       "סכום לפי חוקה — לא נקבע",
                                       "סכום מוזן ידנית", "לא משתתף בחישובים",
                                       "לא מוגדר בחוברת")
            if cnt.get(k)))
        sub.font = Font(size=10, color=MUTED)
        wsc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        _header_row(wsc, 4, ["סמל", "שם הסמל", "סיווג", "פירוט"], [10, 26, 22, 46])
        for i, x in enumerate(codes_index, start=5):
            for c_i, v in enumerate((x["code"], x["name"], x["kind"], x["note"]), start=1):
                cell = wsc.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
                if c_i == 4:
                    cell.alignment = Alignment(wrap_text=True)
            fill = KIND_COLOR.get(x["kind"])
            if fill:
                for c_i in range(1, 5):
                    wsc.cell(row=i, column=c_i).fill = PatternFill("solid", fgColor=fill)
        wsc.freeze_panes = "A5"
        wsc.auto_filter.ref = f"A4:D{4 + len(codes_index)}"

    # ---- שינויי סטטוס בין חודשים -------------------------------------------------
    flips = compute_flips(per_emp)
    if flips:
        wsf = wb.create_sheet("שינויי סטטוס")
        wsf.sheet_view.rightToLeft = True
        wsf.freeze_panes = "A2"
        _header_row(wsf, 1, ["מסד עובד", "משרד", "דרגה", "תקין בחודש",
                             "שגוי בחודש", "פער ₪"], [12, 22, 8, 12, 12, 12])
        bad = Font(color=BAD_TXT, bold=True)
        for i, f in enumerate(flips, start=2):
            vals = [f["id"], f["ministry"], f["darga"], f["from"], f["to"], f["diff"]]
            for c_i, v in enumerate(vals, start=1):
                cell = wsf.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
            wsf.cell(row=i, column=1).number_format = INT
            wsf.cell(row=i, column=6).number_format = MONEY
            wsf.cell(row=i, column=5).font = bad
            wsf.cell(row=i, column=6).font = bad
        wsf.auto_filter.ref = f"A1:F{len(flips) + 1}"

    # ---- פר עובד ---------------------------------------------------------------
    # One row per employee per file. On a many-file run this sheet is the whole
    # file size — 288k rows is ~307 MB of XML and pushes the workbook past the
    # 30 MB that can be sent to the user — so it can be omitted. Everything
    # else (dashboard, coverage, work queue, month-over-month) is unaffected.
    if not skip_per_employee:
        _emp_sheet(wb, "פר עובד", per_emp, "PerEmployee", highlight_invalid=True)

    # ---- המלצות ל-Progim (מצב "Progim בלבד") ---------------------------------
    if recs:
        CAT_HE = {"shekel_mismatch": "סכום שקלי לא תואם",
                  "unstable_rule": "כלל לא יציב בקובץ",
                  "base_noise": "עיגול-בסיס (~₪15)",
                  "base_precision": "דיוק ותק/דרגה"}
        wsr = wb.create_sheet("המלצות ל-Progim")
        wsr.sheet_view.rightToLeft = True
        wsr.merge_cells("A1:F1")
        t = wsr.cell(row=1, column=1,
                     value="המלצות לעדכון ה-Progim — פערים שהריצה הרגילה מנטרלת, "
                           "לסקירה והחלטה (לא תוקנו אוטומטית בתוכנה)")
        t.font = Font(bold=True, size=12, color=NAVY)
        wsr.row_dimensions[1].height = 22
        _header_row(wsr, 2, ["קטגוריה", "סמל", "רכיב", "מס' עובדים",
                             "סכום פער ₪", "המלצה"], [18, 8, 22, 12, 14, 60])
        wsr.freeze_panes = "A3"
        for i, r in enumerate(recs, start=3):
            wsr.cell(row=i, column=1, value=CAT_HE.get(r["category"], r["category"]))
            wsr.cell(row=i, column=2, value=r["code"])
            wsr.cell(row=i, column=3, value=r["name"])
            wsr.cell(row=i, column=4, value=r["count"]).number_format = INT
            c5 = wsr.cell(row=i, column=5, value=r["sum"]); c5.number_format = MONEY
            c6 = wsr.cell(row=i, column=6, value=r["suggestion"]); c6.alignment = Alignment(wrap_text=True)

    # ---- ריכוז לפי סיבה — the one-page answer for management ------------------
    # Every real error grouped by the component that caused it, so a systemic
    # single-code problem is instantly visible instead of being buried in rows.
    real_rows = [r for r in per_emp if r["err_cat"] == "real"]
    cause = defaultdict(lambda: {"n": 0, "sum": 0.0, "under": 0, "over": 0})
    for r in real_rows:
        keys = set()
        for part in (r["flags"] or "").split(";"):
            code = part.strip().split(" ")[0]
            if code.isdigit():
                keys.add(code)
        if not keys:
            keys = {"בסיס/ותק"}
        for k in keys:
            c = cause[k]
            c["n"] += 1
            c["sum"] += abs(r["total_diff"] or 0.0)
            if (r["total_diff"] or 0) > 0.5:
                c["under"] += 1
            elif (r["total_diff"] or 0) < -0.5:
                c["over"] += 1
    if cause:
        ws5 = wb.create_sheet("ריכוז לפי סיבה")
        ws5.sheet_view.rightToLeft = True
        ws5.merge_cells("A1:G1")
        t5 = ws5.cell(row=1, column=1,
                      value="שגיאות אמת לפי הרכיב הגורם — מה מערכתי ומה נקודתי")
        t5.font = Font(bold=True, size=12, color=NAVY)
        ws5.row_dimensions[1].height = 22
        _header_row(ws5, 2, ["סמל / סיבה", "שם הרכיב", "עובדים",
                             "% מכלל שגיאות האמת", "חשיפה ₪", "שולם בחסר",
                             "שולם ביתר"], [14, 24, 11, 16, 14, 12, 12])
        ws5.freeze_panes = "A3"
        n_real = len(real_rows) or 1
        names = {str(g["code"]): g["name"] for g in (code_gaps or [])}
        for i, (k, c) in enumerate(sorted(cause.items(), key=lambda kv: -kv[1]["n"]),
                                   start=3):
            vals = [k, names.get(k, "שכר בסיס / ותק" if k == "בסיס/ותק" else ""),
                    c["n"], c["n"] / n_real, round(c["sum"], 2), c["under"], c["over"]]
            for c_i, v in enumerate(vals, start=1):
                cell = ws5.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
                if c_i in (3, 6, 7):
                    cell.number_format = INT
                if c_i == 4:
                    cell.number_format = "0.0%"
                    # A single code owning most of the errors is systemic.
                    if v >= 0.5:
                        cell.font = Font(bold=True, color=BAD_TXT)
                if c_i == 5:
                    cell.number_format = MONEY

    # ---- פילוח משרדים ----------------------------------------------------------
    agg = defaultdict(lambda: Counter())
    for r in per_emp:
        agg[r["ministry"] or "—"][r["status"]] += 1
    ws4 = wb.create_sheet("פילוח משרדים")
    ws4.sheet_view.rightToLeft = True
    ws4.freeze_panes = "A2"
    # Real errors + their ₪ per ministry: where the remaining work actually is.
    real_by_m, money_by_m = Counter(), defaultdict(float)
    for r in per_emp:
        if r["err_cat"] == "real":
            m = r["ministry"] or "—"
            real_by_m[m] += 1
            money_by_m[m] += abs(r["total_diff"] or 0.0)
    _header_row(ws4, 1, ["משרד / גוף", "עובדים", "תקין", "שגוי",
                         "% תקינות (פעילים)", "שגיאות אמת", "חשיפה ₪"],
                [26, 11, 11, 9, 16, 12, 14])
    # Sorted by real errors — the ministries that still need work come first.
    rows = sorted(agg.items(),
                  key=lambda kv: (-real_by_m.get(kv[0], 0), -sum(kv[1].values())))
    for i, (name, c) in enumerate(rows, start=2):
        act = c["valid"] + c["invalid"]
        vals = [name, sum(c.values()), c["valid"], c["invalid"],
                (c["valid"] / act) if act else None,
                real_by_m.get(name, 0), round(money_by_m.get(name, 0.0), 2)]
        for c_i, v in enumerate(vals, start=1):
            cell = ws4.cell(row=i, column=c_i, value=v)
            cell.border = THIN_BOX
            if c_i in (2, 3, 4, 6):
                cell.number_format = INT
            if c_i == 4 and c["invalid"]:
                cell.font = Font(color=BAD_TXT, bold=True)
            if c_i == 5:
                cell.number_format = "0.0%"
            if c_i == 6 and v:
                cell.font = Font(color=BAD_TXT, bold=True)
            if c_i == 7:
                cell.number_format = MONEY
    last4 = len(rows) + 1
    ws4.conditional_formatting.add(
        f"B2:B{last4}", DataBarRule(start_type="num", start_value=0,
                                    end_type="max", color=BAR_BLUE, showValue=True))
    rng4 = f"E2:E{last4}"
    ws4.conditional_formatting.add(rng4, CellIsRule(
        operator="greaterThanOrEqual", formula=["0.99"],
        fill=PatternFill("solid", fgColor=GOOD_BG)))
    ws4.conditional_formatting.add(rng4, CellIsRule(
        operator="lessThan", formula=["0.95"],
        fill=PatternFill("solid", fgColor=BAD_BG)))
    tbl4 = Table(displayName="Ministries", ref=f"A1:E{last4}")
    tbl4.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
    ws4.add_table(tbl4)

    # ---- מגמה בין-חודשית: native Excel charts on the dashboard -----------------
    months = sorted({r["month"] for r in summary},
                    key=lambda m: (m.split("/")[-1], m.split("/")[0]))
    if len(months) > 1:
        per_month = {m: Counter() for m in months}
        for r in summary:
            for k in ("workers", "valid", "invalid"):
                per_month[r["month"]][k] += r[k]
        anchor = trow + 3
        ws.cell(row=anchor - 1, column=1, value="מגמה בין-חודשית").font = \
            Font(bold=True, color=NAVY)
        _header_row(ws, anchor, ["חודש", "% תקינות", "שגויים"], [11, 11, 9])
        for i, m in enumerate(months, start=anchor + 1):
            c = per_month[m]
            act = c["valid"] + c["invalid"]
            ws.cell(row=i, column=1, value=m).border = THIN_BOX
            pc = ws.cell(row=i, column=2, value=(c["valid"] / act) if act else None)
            pc.number_format = "0.00%"
            pc.border = THIN_BOX
            ic = ws.cell(row=i, column=3, value=c["invalid"])
            ic.number_format = INT
            ic.border = THIN_BOX
        from openpyxl.chart import BarChart, LineChart, Reference
        lastm = anchor + len(months)
        line = LineChart()
        line.title = "% תקינות לפי חודש"
        line.height, line.width = 7, 13
        line.y_axis.numFmt = "0%"
        line.add_data(Reference(ws, min_col=2, min_row=anchor, max_row=lastm),
                      titles_from_data=True)
        line.set_categories(Reference(ws, min_col=1, min_row=anchor + 1, max_row=lastm))
        ws.add_chart(line, f"E{anchor - 1}")
        bar = BarChart()
        bar.title = "שגויים לפי חודש"
        bar.height, bar.width = 7, 13
        bar.add_data(Reference(ws, min_col=3, min_row=anchor, max_row=lastm),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=anchor + 1, max_row=lastm))
        ws.add_chart(bar, f"E{anchor + 15}")

        # ---- שינויים בין חודשים: עובד תקין שהפך שגוי ----------------------------
        month_idx = {m: i for i, m in enumerate(months)}
        by_worker = defaultdict(dict)
        for r in per_emp:
            if r["status"] in ("valid", "invalid"):
                by_worker[r["worker_id"]][r["month"]] = r
        regressions = []
        for wid, per_m in by_worker.items():
            ms = sorted(per_m, key=lambda m: month_idx.get(m, 99))
            for a, b in zip(ms, ms[1:]):
                if per_m[a]["status"] == "valid" and per_m[b]["status"] == "invalid":
                    cur = per_m[b]
                    regressions.append({
                        "worker_id": wid, "ministry": cur["ministry"],
                        "darga": cur["darga"], "from_m": a, "to_m": b,
                        "total_diff": cur["total_diff"], "flags": cur["flags"],
                        "diag": cur["diag"],
                    })
        ws5 = wb.create_sheet("שינויים בין חודשים")
        ws5.sheet_view.rightToLeft = True
        ws5.freeze_panes = "A3"
        ws5.merge_cells("A1:H1")
        h = ws5["A1"]
        h.value = ("עובדים שהיו תקינים בחודש מוקדם והפכו שגויים בחודש מאוחר — "
                   f"{len(regressions)} מקרים · ממוין לפי גודל הפער")
        h.font = Font(bold=True, color=NAVY)
        _header_row(ws5, 2, ["מסד עובד", "משרד", "דרגה", "תקין בחודש",
                             "שגוי בחודש", "הפרש כולל", "רכיבים חריגים", "אבחון"],
                    [12, 22, 8, 12, 12, 12, 34, 34])
        regressions.sort(key=lambda r: abs(r["total_diff"] or 0), reverse=True)
        for i, r in enumerate(regressions, start=3):
            vals = [r["worker_id"], r["ministry"], r["darga"], r["from_m"],
                    r["to_m"], r["total_diff"], r["flags"], r["diag"]]
            for c_i, v in enumerate(vals, start=1):
                cell = ws5.cell(row=i, column=c_i, value=v)
                cell.border = THIN_BOX
                if c_i == 1:
                    cell.number_format = INT
                if c_i == 6:
                    cell.number_format = MONEY
                    cell.font = Font(color=BAD_TXT, bold=True)

    wb.save(out_path)


def main_cli():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("files", nargs="+", help="גולמי .xlsx files")
    ap.add_argument("--out", default="unified.xlsx")
    ap.add_argument("--no-per-employee", action="store_true",
                    help="לדלג על גיליון 'פר עובד' — מקטין דרמטית קובץ של ריצה רב-חודשית")
    ap.add_argument("--pure", action="store_true",
                    help="הרצת ה-Progim כפי-שהוא (ללא תוספות התוכנה) + גיליון המלצות")
    args = ap.parse_args()
    t0 = time.time()
    summary, per_emp, code_gaps, recs, uncovered, codes_index = collect(
        args.files, pure=args.pure)
    print(f"עיבוד: {time.time() - t0:.0f}ש · כותב workbook ({len(per_emp):,} שורות)...")
    write_workbook(summary, per_emp, args.out, code_gaps,
                   recs=recs if args.pure else None, uncovered=uncovered,
                   codes_index=codes_index,
                   skip_per_employee=args.no_per_employee)
    inv = sum(1 for r in per_emp if r["status"] == "invalid")
    mode = "Progim בלבד" if args.pure else "רגיל"
    print(f"נכתב: {args.out} · {len(per_emp):,} רשומות · {inv:,} שגויים · מצב {mode}")
    if args.pure:
        print(f"המלצות ל-Progim: {len(recs)} שורות")


if __name__ == "__main__":
    main_cli()
