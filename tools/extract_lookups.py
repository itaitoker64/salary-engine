"""
extract_lookups.py — build lookups.json from a Progim engine workbook.

The Progim_*.xlsm file is the authoritative Israeli civil-service pay engine.
This script extracts the three tables the salary simulator needs into a compact,
version-controllable JSON so the API doesn't depend on the multi-MB workbook at
runtime:

  - darga:    {grade_label -> base salary at seniority 0}        (DARGA sheet)
  - vetek:    {track_code -> {seniority_years -> multiplier}}     (vetek sheet)
  - track_max:{track_code -> max seniority years}                 (DERUG sheet)
  - tracks:   {track_code -> track name}                          (DERUG sheet)

Usage:
    python tools/extract_lookups.py path/to/Progim.xlsm [lookups.json]
"""
import json
import sys
from pathlib import Path

import openpyxl


def extract(xlsm_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)

    # DARGA: col B = grade label, col D = base salary at seniority 0
    darga: dict[str, float] = {}
    for row in wb["DARGA"].iter_rows(min_row=2, values_only=True):
        label, base = row[1], row[3]
        if label is not None and base is not None:
            try:
                darga[str(label).strip()] = float(base)
            except (TypeError, ValueError):
                pass

    # vetek: row 1 holds the track NAME per column, row 2 the track CODE (the
    # value the payroll "קוד דרוג" uses), col A = seniority years. Only columns
    # that carry BOTH a name (row 1) and an integer code (row 2) are real track
    # columns — this skips the unrelated "וותק חובה"/"דרוג" mini-tables to the
    # right, which otherwise clobber the mapping.
    rows = list(wb["vetek"].iter_rows(min_row=1, values_only=True))
    name_row, track_row = rows[0], rows[1]
    col_for_track: dict[int, int] = {}
    for ci in range(min(len(name_row), len(track_row))):
        name, code = name_row[ci], track_row[ci]
        if name is None or not str(name).strip() or code is None:
            continue
        try:
            col_for_track[int(code)] = ci
        except (TypeError, ValueError):
            continue
    vetek: dict[int, dict[float, float]] = {}
    for track, ci in col_for_track.items():
        table: dict[float, float] = {}
        for r in rows[2:]:
            years = r[0]
            mult = r[ci] if ci < len(r) else None
            if years is None or mult is None:
                continue
            try:
                table[float(years)] = float(mult)
            except (TypeError, ValueError):
                pass
        if table:
            vetek[track] = table

    # DERUG: track code (col D, "קוד דרוג חיל\"ן") -> name (col B) and max vatek (col E)
    track_max: dict[int, float] = {}
    tracks: dict[int, str] = {}
    for row in wb["DERUG"].iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            break
        name, code, vmax = row[1], row[3], row[4]
        if code is None:
            continue
        try:
            code = int(code)
        except (TypeError, ValueError):
            continue
        if name is not None:
            tracks[code] = str(name).strip()
        if vmax is not None:
            try:
                track_max[code] = float(vmax)
            except (TypeError, ValueError):
                pass
    wb.close()

    return {
        "darga": darga,
        "vetek": {str(k): v for k, v in vetek.items()},
        "track_max": {str(k): v for k, v in track_max.items()},
        "tracks": {str(k): v for k, v in tracks.items()},
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "lookups.json"
    data = extract(src)
    Path(out).write_text(json.dumps(data, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"Wrote {out}: {len(data['darga'])} grades, "
          f"{len(data['vetek'])} tracks, {len(data['track_max'])} capped tracks")
