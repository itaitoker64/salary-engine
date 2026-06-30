"""
extract_ministries.py — build ministries.json: the list of ministries/units
(code → name) that appear in a גולמי file, with a worker count each.

Using the codes that actually occur in the payroll data guarantees the dropdown
in the UI matches what users validate against.

Usage:
    python tools/extract_ministries.py path/to/golmi.xlsx ministries.json
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


def extract(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["גולמי"] if "גולמי" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next((i for i, r in enumerate(rows[:8])
                     if r and any(str(c).strip() in ("קוד משרד/גוף", "קוד משרד") for c in r if c)), 0)
    header = [str(c).strip() if c is not None else "" for c in rows[header_i]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    ci_code = col("קוד משרד/גוף", "קוד משרד")
    ci_name = col("משרד/גוף", "שם משרד")
    ci_worker = col("קוד מסב", "מסד")

    seen = defaultdict(lambda: {"name": "", "workers": set()})
    for r in rows[header_i + 1:]:
        if ci_code is None or ci_code >= len(r) or r[ci_code] is None:
            continue
        code = int(r[ci_code])
        m = seen[code]
        if ci_name is not None and r[ci_name]:
            m["name"] = str(r[ci_name]).strip()
        if ci_worker is not None and ci_worker < len(r) and r[ci_worker] is not None:
            m["workers"].add(r[ci_worker])

    ministries = [{"code": code, "name": m["name"], "workers": len(m["workers"])}
                  for code, m in seen.items()]
    ministries.sort(key=lambda x: -x["workers"])
    wb.close()
    return {"ministries": ministries}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    out = sys.argv[2] if len(sys.argv) > 2 else "ministries.json"
    data = extract(sys.argv[1])
    Path(out).write_text(json.dumps(data, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"Wrote {out}: {len(data['ministries'])} ministries")
